#!/usr/bin/env python3
"""
Redfin + Zillow + DBEDT → Housing Affordability Tracker price updater.

Downloads:
  - Redfin public S3 TSV files → median sale prices (SFH + condo) per county
  - Zillow ZORI county CSV     → median asking rent per county
  - DBEDT QSER construction XLSX → private building authorization values per county
  - HHFDC county income schedule PDFs → HUD median family income per county
  - HUD State Income Limits PDF  → statewide median family income

Patches squarespace-single-file.html in-place with fresh values.

No API keys needed — all sources are public.
Redfin data: monthly (Friday of the third full week).
Zillow ZORI: monthly.
DBEDT QSER:  quarterly.

Usage:
    python3 redfin-price-updater.py                   # update squarespace-single-file.html
    python3 redfin-price-updater.py --dry-run          # print changes without writing
    python3 redfin-price-updater.py --file other.html  # target a different file

Sources:
  Redfin:  https://www.redfin.com/news/data-center/
  Zillow:  https://www.zillow.com/research/data/
  DBEDT:   https://dbedt.hawaii.gov/economic/qser/construction/
"""

import csv
import gzip
import io
import os
import re
import statistics
import sys
from pathlib import Path

# Shared HTTP helper — adds timeout + retry to all fetches
sys.path.insert(0, str(Path(__file__).resolve().parent))
from ha_common.http_client import fetch_bytes, fetch_text   # noqa: E402
from ha_common.html_patcher import patch_html_files  # noqa: E402

try:
    import openpyxl
    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False

try:
    import pdfplumber
    _PDFPLUMBER_AVAILABLE = True
except ImportError:
    _PDFPLUMBER_AVAILABLE = False

# ─── CONFIG ─────────────────────────────────────────────────────
STATE_URL  = "https://redfin-public-data.s3.us-west-2.amazonaws.com/redfin_market_tracker/state_market_tracker.tsv000.gz"
COUNTY_URL = "https://redfin-public-data.s3.us-west-2.amazonaws.com/redfin_market_tracker/county_market_tracker.tsv000.gz"
ZORI_URL   = "https://files.zillowstatic.com/research/public_csvs/zori/County_zori_uc_sfrcondomfr_sm_month.csv"
DBEDT_URL  = "https://files.hawaii.gov/dbedt/economic/data_reports/qser/E-construction-tables.xlsx"

# HHFDC county income schedule PDFs (HUD income limits, published by state of Hawaii).
# The "MEDIAN" column in each schedule is HUD's FY2025 4-person median family income.
HHFDC_PDF_TEMPLATE = "https://dbedt.hawaii.gov/hhfdc/files/2025/05/{county}-County-2025.pdf"
HHFDC_COUNTIES     = ["Honolulu", "Hawaii", "Maui", "Kauai"]

# HUD State Income Limits report (FY2025) — contains each state's MFI including HI.
HUD_STATE_IL_URL   = "https://www.huduser.gov/portal/datasets/il/il25/State-Incomelimits-Report-FY25.pdf"
HUD_FY             = "FY 2025"

# DBEDT E-8 column header → countyData key (columns in order: State, Honolulu, Hawaii, Kauai, Maui)
# The header row in the sheet uses newlines inside cell values
DBEDT_COL_KEYS = ["State", "Honolulu", "Hawaii", "Kauai", "Maui"]  # columns 1–5 in E-8

# ─── Affordability assumptions (mirror the dashboard JS) ─────────
# These reproduce calcAffordPrice()/mcardV2() in index.html so the price-
# derived metrics (sfhIdx/sfhGap/sfhMortgage/sfhPTI + condo equivalents)
# are recomputed each run instead of drifting as stale literals.
#
# MORTGAGE_RATE_PCT MUST stay in sync with the rate-slider DEFAULT in the
# HTML (id="rate-slider" value=…). At the default slider position the live
# banner's affordPrice and the stored idx/gap agree; if they diverge the
# stored metrics would contradict the calculator on first paint. Bump both
# together when refreshing the Freddie Mac PMMS rate.
MORTGAGE_RATE_PCT   = 6.38   # 30-yr fixed, Freddie Mac PMMS (matches slider default)
DOWN_PAYMENT_FRAC   = 0.20   # 20% down → LTV 0.80
DTI_FRONT_FRAC      = 0.30   # 30% of gross income to P&I
MORTGAGE_TERM_MONTHS = 360   # 30-year amortization

# ─── Rent anchor year (SINGLE SOURCE OF TRUTH) ──────────────────
# Both the ACS contract-rent dollar anchor and the BLS rent-CPI base-year
# average must align on the same vintage — otherwise the scaling factor
# "BLS(now) / BLS(anchor_year_avg)" applied to "ACS(anchor_year) dollars"
# produces a dollar value that is anchored to a different year than the
# index says. Keep both pointing at the same YEAR constant.
#
# RE-ANCHORING CADENCE (see METHODOLOGY.md): bump this every December
# when a new ACS 5-year vintage is released. Pull the fresh Honolulu
# contract rent directly from the Census API response — no more
# hardcoded dollar values.
RENT_ANCHOR_YEAR = "2024"

# Census ACS — contract rent (B25058_001E, utilities excluded)
CENSUS_ACS_YEAR = RENT_ANCHOR_YEAR
CENSUS_BASE_URL = f"https://api.census.gov/data/{CENSUS_ACS_YEAR}/acs/acs5"
CENSUS_RENT_VAR = "B25058_001E"   # median contract rent (no utilities) — comparable to Zillow ZORI
CENSUS_NAME_MAP = {
    "Honolulu County, Hawaii": "Honolulu",
    "Hawaii County, Hawaii":   "Hawaii",
    "Maui County, Hawaii":     "Maui",
    "Kauai County, Hawaii":    "Kauai",
}

# Census API key (env var). Required for ACS5 calls — Census tightened
# the anonymous policy in 2025; the previous unauthenticated path now
# returns "Missing Key" HTML. Sign up free at
# https://api.census.gov/data/key_signup.html and set as CENSUS_API_KEY
# in GitHub Actions + local env.
CENSUS_API_KEY = os.environ.get("CENSUS_API_KEY", "")

# Realized-burden anchor: ACS B25071 (median GRAPI for renters) +
# B25092 (median SMOCAPI for owners with mortgage) + B25070/B25091
# cost-burden distributions. Pulled from the same 5-year vintage as
# the contract-rent anchor (RENT_ANCHOR_YEAR) so the dashboard's
# "rent" and "burden" numbers describe the same population window.
CENSUS_BURDEN_VARS = [
    "B25071_001E",   # Median GRAPI (renter households, %)
    "B25092_002E",   # Median SMOCAPI for owner-occupied with mortgage (%)
    # B25070 — Gross rent as % of household income (renter distribution)
    "B25070_001E",   # Total renters
    "B25070_007E",   # 30.0 – 34.9 %
    "B25070_008E",   # 35.0 – 39.9 %
    "B25070_009E",   # 40.0 – 49.9 %
    "B25070_010E",   # 50.0 %+
    "B25070_011E",   # Not computed
    # B25091 — Mortgage status × SMOCAPI (owner distribution, with-mortgage rows)
    "B25091_002E",   # With mortgage (total)
    "B25091_006E",   # With mortgage, 30.0 – 34.9 %
    "B25091_007E",   # With mortgage, 35.0 – 39.9 %
    "B25091_008E",   # With mortgage, 40.0 – 49.9 %
    "B25091_009E",   # With mortgage, 50.0 %+
    "B25091_010E",   # With mortgage, not computed
]

# BLS series for nowcasting from the ACS 5-year mid-point to the
# current period. Rent uses Honolulu rent CPI (already fetched
# elsewhere — re-fetched here so the burden pipeline is self-contained).
# Owner SMOCAPI is dominated by sticky locked-in P&I + slow-growing
# tax/ins/util → tracks Honolulu all-items CPI as a stand-in. Income
# uses Hawaii state private avg weekly earnings (CES).
BLS_CPI_ALL_ITEMS_HNL = "CUURS49ASA0"           # Honolulu, all items
BLS_WAGES_HI_PRIVATE  = "SMU15000000500000011"  # Hawaii state private avg weekly earnings (NSA)

# BLS CPI: Honolulu MSA — "Rent of primary residence" (existing tenants, not new leases)
# Series CUURS49ASEHA, not seasonally adjusted, base 1982-84=100.
# We scale the live ACS Honolulu contract rent (fetched each run) by the
# BLS index ratio (latest / anchor-year avg) to get a monthly-current estimate.
BLS_API_URL     = "https://api.bls.gov/publicAPI/v2/timeseries/data/"
BLS_RENT_SERIES = "CUURS49ASEHA"
BLS_BASE_YEAR   = RENT_ANCHOR_YEAR

# NTR/ATR national benchmarks — manually refreshed quarterly from BLS research
# series R-CPI-NTR and R-CPI-ATR. Used only for a dev-facing sanity check on
# the Honolulu rent nowcast (see audit_rent_nowcast_vs_ntr below). Not read by
# the dashboard UI. See METHODOLOGY.md § Quarterly NTR/ATR benchmark refresh.
NTR_ATR_BENCHMARKS_PATH = Path(__file__).parent / "data" / "ntr_atr_benchmarks.json"

# Blended rent nowcast — per-county weights for combining the lagging BLS
# rent CPI with Zillow ZORI's leading asking-rent signal. Both series are
# growth factors vs. RENT_ANCHOR_YEAR applied to the same ACS dollar anchor:
#   blended_rent = acs_anchor × (cpi_w · bls_ratio + (1−cpi_w) · zori_ratio)
#
# The BLS Honolulu rent index (CUURS49ASEHA) is a Honolulu-only sample —
# applying it to neighbor islands at the same 70% weight implies that
# Honolulu rent dynamics are a reliable proxy for Maui/Hawaiʻi/Kauaʻi,
# which is methodologically weak. Outer islands get 50/50 so the only
# county-specific signal (ZORI) gets equal say. Honolulu and State (which
# is Honolulu-dominated by population weight ~72%) keep 70/30.
# Cleveland Fed WP 22-38r motivates the lag-vs-leading blend structure.
BLENDED_RENT_CPI_WEIGHTS = {
    "State":    0.70,   # Honolulu-dominated; CPI is regionally representative
    "Honolulu": 0.70,   # genuine local CPI signal
    "Maui":     0.50,   # ZORI is the only county-specific source
    "Hawaii":   0.50,
    "Kauai":    0.50,
}
# Backward-compat fallback for any caller still passing through the old name
BLENDED_RENT_CPI_WEIGHT = BLENDED_RENT_CPI_WEIGHTS["Honolulu"]

# Per-county 3-leg blend weights — ONLY the two counties where adding a
# genuinely county-specific HUD Fair Market Rent (FMR) leg is justified
# (M4 backtest + 3-leg addendum, docs/rent_nowcast_backtest.md):
#
#   • Hawaiʻi — realized ACS rent +27% (2021→24) outran BOTH Honolulu CPI
#     (+15%) and ZORI (+16%); only HUD FMR (+37%) captured the divergence,
#     cutting the Hawaiʻi backtest MAPE 9.21% → ~4.8%. The CPI leg is
#     Honolulu-only, so without FMR the Big Island has no signal that can
#     diverge from Honolulu — in either direction (surge OR plateau).
#   • Kauaʻi — has NO ZORI history (its ZORI leg is the statewide proxy), so
#     FMR is its first genuinely county-specific signal. CAVEAT: FMR slightly
#     *worsens* Kauaʻi's measured MAPE (status-quo CPI/proxy-ZORI 4.00% vs
#     CPI/FMR 5.74%) because Kauaʻi's realized growth was modest (+14%) while
#     FMR overshot (+25%) over the test window. FMR is therefore added at a
#     MODEST 0.20 weight here — for robustness (a real local leg), not measured
#     accuracy. Keep its weight small and re-validate annually.
#
# Maui (CPI/ZORI 50/50 — ZORI is the better county signal there) and
# Honolulu/State (CPI-led 0.70) are intentionally NOT given an FMR leg.
# Counties absent from this dict fall back to the 2-leg BLENDED_RENT_CPI_WEIGHTS.
# Weights are explicit (cpi/zori/fmr) and must sum to 1.0.
BLENDED_RENT_3LEG_WEIGHTS = {
    "Hawaii": {"cpi": 0.34, "zori": 0.33, "fmr": 0.33},
    "Kauai":  {"cpi": 0.40, "zori": 0.40, "fmr": 0.20},
}

# Number of trailing monthly periods to average when computing the BLS
# rent-CPI ratio that feeds the blended nowcast (Q1 of the rent-nowcast
# improvement plan). ZORI input is already smoothed at the same window;
# matching it here means both legs of the blend reflect a 3-month mean
# instead of a single bumpy print.
BLS_RENT_SMOOTHING_WINDOW = 3

# Per-county ACS vintage override. Default is the 5-year (CENSUS_BASE_URL).
# Maui uses the 1-year because the 2020–2024 5-yr dilutes the post-Lahaina
# rental shock by averaging it with pre-fire years; the 1-yr 2024 captures
# the single-year impact. ACS 1-yr only publishes for areas ≥65k pop.
# Other counties qualify but currently keep 5-yr for stability.
COUNTY_ANCHOR_OVERRIDE = {
    "Maui": "acs1",
}

# Redfin region name → countyData key in the HTML file
COUNTY_MAP = {
    "Honolulu County, HI": "Honolulu",
    "Hawaii County, HI":   "Hawaii",
    "Maui County, HI":     "Maui",
    "Kauai County, HI":    "Kauai",
}

# Zillow ZORI RegionName → countyData key
ZORI_COUNTY_MAP = {
    "Honolulu County": "Honolulu",
    "Hawaii County":   "Hawaii",
    "Maui County":     "Maui",
    "Kauai County":    "Kauai",
}

# Redfin property type → which countyData field to update
PROP_TYPE_MAP = {
    "Single Family Residential": "sfhPrice",
    "Condo/Co-op":               "condoPrice",
}

DEFAULT_FILES = [
    Path(__file__).parent / "squarespace-single-file.html",
    Path(__file__).parent / "index.html",
]
# ────────────────────────────────────────────────────────────────


def download_tsv(url: str) -> list[dict]:
    """Download a gzipped TSV from Redfin's S3 bucket and return rows as dicts."""
    print(f"  Downloading {url.split('/')[-1]}...")
    raw = gzip.decompress(fetch_bytes(url))
    reader = csv.DictReader(io.StringIO(raw.decode("utf-8")), delimiter="\t")
    return list(reader)


SALE_PRICE_SMOOTHING_WINDOW = 3   # months for trailing median — see comment below


def extract_hawaii_prices(rows: list[dict], region_col: str, region_values: dict) -> dict:
    """
    Filter rows to Hawaii regions + target property types and return the
    SALE_PRICE_SMOOTHING_WINDOW-month trailing **median** sale price for each
    (region, property_type): {countyData_key: {sfhPrice: int, condoPrice: int}}.

    Why median, not the single latest month: Redfin reports the median sale
    price of whatever closed that month, and thin Hawaiʻi submarkets transact
    in tiny volumes (Kauaʻi SFH ≈ 27 sales/mo, Hawaiʻi condos ≈ 33). A single
    luxury batch swings the headline ±20% — e.g. the latest Kauaʻi SFH print
    sat ~19% above its own 3-month mean. Taking the median of the last three
    monthly prints damps that sampling noise while staying robust to a single
    outlier month (a mean would let one $5M sale drag the figure). This mirrors
    the ZORI trailing-mean treatment in fetch_zori_asking_rents().

    Redfin market-tracker files are monthly (PERIOD_DURATION == 30), so each
    (region, type, month) is one row; we still pin to the latest row's duration
    defensively in case Redfin ever mixes cadences into the same export.
    """
    # Filter to Hawaii + relevant property types
    filtered = []
    for row in rows:
        region = row.get(region_col, "").strip('"')
        prop   = row.get("PROPERTY_TYPE", "").strip('"')
        price  = row.get("MEDIAN_SALE_PRICE", "").strip('"')
        period = row.get("PERIOD_BEGIN", "").strip('"')
        dur    = row.get("PERIOD_DURATION", "").strip('"')

        if region not in region_values or prop not in PROP_TYPE_MAP:
            continue
        if not price or not period:
            continue

        filtered.append({
            "key":    region_values[region],
            "field":  PROP_TYPE_MAP[prop],
            "price":  int(float(price)),
            "period": period,
            "dur":    dur,
        })

    # Group every observation per (key, field), then take the trailing-median
    # of the most recent N monthly prints (restricted to the latest row's
    # cadence, deduped to one price per month).
    series: dict[tuple[str, str], list[dict]] = {}
    for row in filtered:
        series.setdefault((row["key"], row["field"]), []).append(row)

    result: dict[str, dict] = {}
    for (key, field), obs in series.items():
        obs.sort(key=lambda r: r["period"])
        latest_dur = obs[-1]["dur"]
        by_period = {r["period"]: r["price"] for r in obs if r["dur"] == latest_dur}
        window_periods = sorted(by_period)[-SALE_PRICE_SMOOTHING_WINDOW:]
        window_prices  = [by_period[p] for p in window_periods]
        smoothed       = int(round(statistics.median(window_prices)))
        latest_period  = window_periods[-1]

        raw_latest = by_period[latest_period]
        delta = (smoothed - raw_latest) / raw_latest * 100 if raw_latest else 0.0
        print(f"  {key:<9} {field:<10} {SALE_PRICE_SMOOTHING_WINDOW}-mo median "
              f"${smoothed:>9,}  (latest ${raw_latest:>9,}, {delta:+.1f}%, "
              f"n={len(window_prices)})")

        if key not in result:
            result[key] = {"period": latest_period}
        result[key][field] = smoothed
        # Keep the most recent period across both property types
        if latest_period > result[key]["period"]:
            result[key]["period"] = latest_period

    return result


ZORI_SMOOTHING_WINDOW = 3   # months for trailing mean — see comment below

def fetch_zori_asking_rents() -> dict:
    """
    Download Zillow ZORI county CSV and extract:
      - A ZORI_SMOOTHING_WINDOW-month trailing mean of asking rent per
        Hawaiʻi county, written as result[key] and used as the displayed
        askRent (and as numerator in the ZORI growth ratio feeding the
        blended rent nowcast).
      - The RENT_ANCHOR_YEAR annual average per county, used as a common
        anchor with BLS rent CPI for the blended nowcast
        (→ result["_anchor_avg"][key])

    Returns {countyData_key: askRent_int, "_period": "YYYY-MM",
             "_anchor_avg": {...}, "_anchor_year": "YYYY", "_yoy_pct": {...},
             "_smoothing_window": int}.
    State-level askRent is derived as a population-weighted average
    (Honolulu ~72%, Hawaii ~14%, Maui ~10%, Kauai ~4%) of the smoothed
    county values.

    Why smooth: ZORI is already a "smoothed mean" (the _sm_ in the URL) but
    its sample size on thin markets — Kauaʻi especially — is small enough
    that a single luxury batch entering the listings pool can swing the
    headline number by 15-20% in one month (e.g. the 2026-04 print of
    $5,255 vs. a 12-month band of $4.3k–$4.5k). Averaging the latest 3
    monthly prints dampens these sampling artifacts without materially
    changing the trend signal Zillow already publishes.

    The anchor-year average MUST track RENT_ANCHOR_YEAR — the BLS rent CPI
    ratio is computed against the same year, and the blended nowcast assumes
    both series share the anchor. A drifted anchor here would silently shift
    the ZORI growth factor relative to BLS. (Anchor stays a full-year mean —
    only the *current* read is windowed.)
    """
    print(f"  Downloading {ZORI_URL.split('/')[-1]}...")
    raw = fetch_text(ZORI_URL)

    reader = csv.reader(io.StringIO(raw))
    headers = next(reader)

    # Pre-compute which column indices belong to RENT_ANCHOR_YEAR for the
    # anchor-avg calc. ZORI column headers are ISO dates like "2024-01-31".
    anchor_prefix = f"{RENT_ANCHOR_YEAR}-"
    cols_anchor = [i for i, h in enumerate(headers) if h.startswith(anchor_prefix)]

    result = {}
    anchor_avg = {}
    yoy_pct = {}   # per-county YoY % using same-month-prior-year column
    latest_date_header = None  # e.g. "2026-03-31" → we'll convert to "2026-03"
    for row in reader:
        if len(row) < 10:
            continue
        region_name = row[2]
        state       = row[5]
        if state != "HI" or region_name not in ZORI_COUNTY_MAP:
            continue

        # Find the last non-empty column (most recent month) — return both value and header
        last_idx = next(
            (i for i in range(len(row) - 1, 8, -1) if row[i].strip()),
            None,
        )
        if last_idx is None:
            continue

        key = ZORI_COUNTY_MAP[region_name]

        # Walk backward from the latest non-empty cell collecting up to
        # ZORI_SMOOTHING_WINDOW numeric values, then take the mean. Skips
        # blank or unparseable cells. Falls back to whatever fewer values
        # are available (handles Zillow's late-arriving small-market
        # coverage — e.g. Kauaʻi has only had data since Feb 2025).
        window_vals = []
        for i in range(last_idx, 8, -1):
            cell = row[i].strip() if i < len(row) else ""
            if not cell:
                continue
            try:
                window_vals.append(float(cell))
            except ValueError:
                continue
            if len(window_vals) >= ZORI_SMOOTHING_WINDOW:
                break
        if not window_vals:
            continue
        result[key] = round(sum(window_vals) / len(window_vals))

        # Capture the column header (date) once; should be identical across counties
        if latest_date_header is None and last_idx < len(headers):
            latest_date_header = headers[last_idx]

        # Anchor-year annual average — skip empty cells / unparseable values
        vals_anchor = []
        for i in cols_anchor:
            if i < len(row) and row[i].strip():
                try:
                    vals_anchor.append(float(row[i]))
                except ValueError:
                    pass
        if vals_anchor:
            anchor_avg[key] = sum(vals_anchor) / len(vals_anchor)

        # YoY: current column vs the column 12 months earlier. ZORI publishes
        # every month so the same position back by 12 is the same calendar month
        # a year ago. Used by the NTR/ATR sanity-check audit.
        if last_idx >= 12 + 9:  # +9 is the first data column (after metadata)
            prior_cell = row[last_idx - 12].strip() if last_idx - 12 < len(row) else ""
            if prior_cell:
                try:
                    prior_val = float(prior_cell)
                    if prior_val > 0:
                        yoy_pct[key] = (float(row[last_idx]) / prior_val - 1.0) * 100.0
                except ValueError:
                    pass

    # Compute statewide weighted average if all counties present
    weights = {"Honolulu": 0.72, "Hawaii": 0.14, "Maui": 0.10, "Kauai": 0.04}
    if all(k in result for k in weights):
        state_avg = sum(result[k] * w for k, w in weights.items())
        result["State"] = round(state_avg)
    # Statewide ZORI YoY via the same population weights, on the counties
    # that actually have a YoY value (Zillow can lag publishing small-market
    # YoYs by a few months; re-normalize across whatever's present).
    present_yoy = {k: weights[k] for k in weights if k in yoy_pct}
    if len(present_yoy) >= 2:
        wsum = sum(present_yoy.values())
        yoy_pct["State"] = sum(
            yoy_pct[k] * (w / wsum) for k, w in present_yoy.items()
        )
    # For the anchor-year average, allow partial coverage (Zillow started
    # publishing some small-market counties like Kauai only recently, so Kauai
    # can be missing from the anchor year even when its current value is
    # reported). When that happens, re-normalize the weights across the
    # counties we actually have.
    present_anchor = {k: weights[k] for k in weights if k in anchor_avg}
    if len(present_anchor) >= 2:
        wsum = sum(present_anchor.values())
        anchor_avg["State"] = sum(
            anchor_avg[k] * (w / wsum) for k, w in present_anchor.items()
        )

    # Convert "2026-03-31" → "2026-03" for consistency with other period fields
    if latest_date_header and len(latest_date_header) >= 7:
        result["_period"] = latest_date_header[:7]

    result["_anchor_avg"] = anchor_avg
    result["_anchor_year"] = RENT_ANCHOR_YEAR
    result["_yoy_pct"] = yoy_pct  # per-county YoY % — used by NTR/ATR audit
    result["_smoothing_window"] = ZORI_SMOOTHING_WINDOW
    return result


def _census_base_url(variant: str = "acs5", year: str = CENSUS_ACS_YEAR) -> str:
    """Build the Census ACS endpoint URL for a given variant (acs5 / acs1).

    The vintage YEAR stays the same — Census publishes both 1-yr and 5-yr
    estimates dated to the same end year. Variant just toggles the sample
    window (1-yr is more responsive; 5-yr is smoother and covers smaller
    geographies). 1-yr is only published for areas with ≥65k population.
    """
    return f"https://api.census.gov/data/{year}/acs/{variant}"


def fetch_census_rent() -> dict:
    """
    Download median contract rent (B25058_001E) for Hawaii state + 4 counties.

    The default vintage is ACS 5-year (CENSUS_BASE_URL). Per-county overrides
    in COUNTY_ANCHOR_OVERRIDE can switch a single county to ACS 1-year — used
    for Maui (post-Lahaina-fire single-year shock that the 5-yr dilutes by
    averaging with pre-fire years). Each override is fetched independently;
    failure falls back to the 5-yr value already in the result dict.

    Contract rent excludes utilities — directly comparable to Zillow ZORI.
    Returns {countyKey: {rent: int, rentAnchorVariant: "acs5"|"acs1"}}
    plus '_year' metadata.
    """
    import json

    def _get(url):
        return json.loads(fetch_bytes(url))

    key_qs = f"&key={CENSUS_API_KEY}" if CENSUS_API_KEY else ""

    # ── 1. Base 5-yr pull for everyone (state + all 4 counties) ─────────
    base_5yr   = _census_base_url("acs5")
    state_url  = f"{base_5yr}?get={CENSUS_RENT_VAR}&for=state:15{key_qs}"
    county_url = f"{base_5yr}?get={CENSUS_RENT_VAR},NAME&for=county:*&in=state:15{key_qs}"

    print(f"  Fetching Census ACS 5-yr {CENSUS_ACS_YEAR} contract rent (B25058_001E)...")
    state_data  = _get(state_url)
    county_data = _get(county_url)

    result = {"_year": CENSUS_ACS_YEAR}

    # State row: [header, data_row]
    s_hdr, s_row = state_data[0], state_data[1]
    result["State"] = {"rent": int(s_row[s_hdr.index(CENSUS_RENT_VAR)]),
                       "rentAnchorVariant": "acs5"}

    # County rows
    c_hdr, *c_rows = county_data
    rent_idx = c_hdr.index(CENSUS_RENT_VAR)
    name_idx = c_hdr.index("NAME")
    # Hawaii FIPS county codes for the 1-yr per-county query (county:* with
    # acs1 returns only ≥65k counties, but for explicit per-county fetches
    # we use the numeric code).
    county_fips = {"Honolulu": "003", "Hawaii": "001", "Maui": "009", "Kauai": "007"}
    for row in c_rows:
        key = CENSUS_NAME_MAP.get(row[name_idx])
        if key:
            result[key] = {"rent": int(row[rent_idx]), "rentAnchorVariant": "acs5"}

    # ── 2. Per-county 1-yr overrides (e.g. Maui post-fire) ──────────────
    for cty, variant in COUNTY_ANCHOR_OVERRIDE.items():
        if variant != "acs1" or cty not in county_fips:
            continue
        fips = county_fips[cty]
        url = (f"{_census_base_url('acs1')}?get={CENSUS_RENT_VAR},NAME"
               f"&for=county:{fips}&in=state:15{key_qs}")
        try:
            data = _get(url)
            hdr, *rows = data
            if rows:
                r_idx = hdr.index(CENSUS_RENT_VAR)
                v_raw = rows[0][r_idx]
                v = int(v_raw) if v_raw and v_raw not in ("-", "null") and int(v_raw) > 0 else None
                if v is not None:
                    result.setdefault(cty, {})
                    prev = result[cty].get("rent")
                    result[cty]["rent"] = v
                    result[cty]["rentAnchorVariant"] = "acs1"
                    pf = f" (was 5-yr ${prev:,})" if prev else ""
                    print(f"  → {cty}: ACS 1-yr {CENSUS_ACS_YEAR} override ${v:,}{pf}")
                else:
                    print(f"  WARNING: {cty} 1-yr override returned suppressed value; "
                          f"falling back to 5-yr ${result.get(cty, {}).get('rent', '?')}")
        except Exception as e:
            print(f"  WARNING: {cty} 1-yr override fetch failed ({e}); "
                  f"keeping 5-yr ${result.get(cty, {}).get('rent', '?')}")

    return result


def fetch_census_bedroom_rent() -> dict:
    """
    Fetch median gross rent by bedroom count (ACS 5-year table B25031
    "Median Gross Rent by Bedrooms"). Direct medians, one variable per
    bedroom bucket:

      B25031_001E — Total median gross rent (all units)
      B25031_002E — No bedroom (studio / 0 BR)
      B25031_003E — 1 bedroom
      B25031_004E — 2 bedrooms
      B25031_005E — 3 bedrooms
      B25031_006E — 4 bedrooms
      B25031_007E — 5+ bedrooms

    Gross rent INCLUDES utilities (different from B25058 contract rent).
    We collapse 3/4/5+ into a single "3+ BR" tile since 4-BR and 5+-BR
    cells are often suppressed or have wide MoEs in smaller counties.

    Returns {countyKey: {bedroomRent: {br0, br1, br2, br3plus}}} plus
    "_year" metadata. Missing/suppressed cells become None and are
    rendered as "—" in the UI. Silent failure when CENSUS_API_KEY is
    unset — the field already has a graceful "no data" fallback.
    """
    import json
    if not CENSUS_API_KEY:
        print(f"  WARNING: CENSUS_API_KEY not set; skipping bedroom rent fetch.")
        return {}

    bedroom_vars = ["B25031_002E", "B25031_003E", "B25031_004E",
                    "B25031_005E", "B25031_006E", "B25031_007E"]
    vars_csv = ",".join(bedroom_vars)
    state_url  = f"{_census_base_url('acs5')}?get=NAME,{vars_csv}&for=state:15&key={CENSUS_API_KEY}"
    county_url = f"{_census_base_url('acs5')}?get=NAME,{vars_csv}&for=county:*&in=state:15&key={CENSUS_API_KEY}"

    print(f"  Fetching Census ACS 5-yr {CENSUS_ACS_YEAR} bedroom rent (B25031)...")
    try:
        state_data  = json.loads(fetch_bytes(state_url))
        county_data = json.loads(fetch_bytes(county_url))
    except (OSError, json.JSONDecodeError) as e:
        print(f"  WARNING: bedroom rent fetch failed ({e}); skipping.")
        return {}

    def _val(row, hdr, name):
        raw = row[hdr.index(name)]
        try:
            v = float(raw)
            # ACS sentinel for "estimate not displayed" or suppressed
            return None if v <= 0 or v <= -666666666 else int(round(v))
        except (TypeError, ValueError):
            return None

    def _build(hdr, row):
        br0 = _val(row, hdr, "B25031_002E")
        br1 = _val(row, hdr, "B25031_003E")
        br2 = _val(row, hdr, "B25031_004E")
        # Collapse 3/4/5+ into one tile, averaging non-null medians
        # (smaller counties suppress 4-BR / 5+-BR).
        br3p_vals = [v for v in (
            _val(row, hdr, "B25031_005E"),
            _val(row, hdr, "B25031_006E"),
            _val(row, hdr, "B25031_007E"),
        ) if v is not None]
        br3plus = int(round(sum(br3p_vals) / len(br3p_vals))) if br3p_vals else None
        return {"br0": br0, "br1": br1, "br2": br2, "br3plus": br3plus}

    result = {"_year": CENSUS_ACS_YEAR}
    s_hdr, s_row = state_data[0], state_data[1]
    result["State"] = {"bedroomRent": _build(s_hdr, s_row)}

    c_hdr, *c_rows = county_data
    name_idx = c_hdr.index("NAME")
    for row in c_rows:
        key = CENSUS_NAME_MAP.get(row[name_idx])
        if key:
            result[key] = {"bedroomRent": _build(c_hdr, row)}

    for key in ("State", "Honolulu", "Maui", "Hawaii", "Kauai"):
        br = result.get(key, {}).get("bedroomRent")
        if br:
            parts = [f"{lbl}=${v:,}" if v else f"{lbl}=—"
                     for lbl, v in (("0BR", br["br0"]), ("1BR", br["br1"]),
                                    ("2BR", br["br2"]), ("3+BR", br["br3plus"]))]
            print(f"  {key:<9} {' · '.join(parts)}")

    return result


def fetch_census_burden_anchor() -> dict:
    """
    Fetch the ACS 5-year realized-burden anchor for Hawaiʻi state + 4 counties.

    Pulls B25071 (median GRAPI for renters), B25092 (median SMOCAPI for
    owners with a mortgage), and the cost-burden distributions B25070
    (renters) and B25091 (owners-with-mortgage). Returns a flat dict
    keyed by countyKey with per-county derived metrics. Cost-burden
    shares are computed by collapsing the 30-34.9/35-39.9/40-49.9/≥50
    buckets into "≥30 %" and the ≥50 bucket into "severely burdened",
    excluding the "not computed" pool from the denominator.

    Returns {countyKey: {tenantGRAPI, ownerSMOCAPI,
                         rentBurdenedPct, rentSeverelyBurdenedPct,
                         ownerBurdenedPct, ownerSeverelyBurdenedPct}}
    plus "_year" metadata. Percentages are decimals (0.0–1.0).

    The Census API requires a key as of mid-2025 — silently returns {}
    if CENSUS_API_KEY isn't set so the rest of the pipeline doesn't fail.
    """
    import json
    if not CENSUS_API_KEY:
        print(f"  WARNING: CENSUS_API_KEY not set; skipping realized-burden anchor fetch.")
        return {}

    vars_csv = ",".join(CENSUS_BURDEN_VARS)
    state_url  = f"{CENSUS_BASE_URL}?get=NAME,{vars_csv}&for=state:15&key={CENSUS_API_KEY}"
    county_url = f"{CENSUS_BASE_URL}?get=NAME,{vars_csv}&for=county:*&in=state:15&key={CENSUS_API_KEY}"

    print(f"  Fetching Census ACS {CENSUS_ACS_YEAR} realized-burden anchor (B25071, B25092, B25070, B25091)...")
    try:
        state_data  = json.loads(fetch_bytes(state_url))
        county_data = json.loads(fetch_bytes(county_url))
    except (OSError, json.JSONDecodeError) as e:
        print(f"  WARNING: realized-burden anchor fetch failed ({e}); skipping.")
        return {}

    def _build_row(hdr, row):
        """Compute the 6 dashboard fields from a single census API row."""
        def fv(name, scale=1.0):
            raw = row[hdr.index(name)]
            v = float(raw)
            # ACS sentinel for "estimate not displayed" (Kalawao, etc.)
            return None if v <= -666666666 else v * scale
        graphi   = fv("B25071_001E", 0.01)   # already a %, store as decimal
        smocapi  = fv("B25092_002E", 0.01)
        r_total  = fv("B25070_001E") or 0
        r_notc   = fv("B25070_011E") or 0
        r_b30_34 = fv("B25070_007E") or 0
        r_b35_39 = fv("B25070_008E") or 0
        r_b40_49 = fv("B25070_009E") or 0
        r_b50p   = fv("B25070_010E") or 0
        r_denom  = max(0, r_total - r_notc)
        rent_burdened     = (r_b30_34 + r_b35_39 + r_b40_49 + r_b50p) / r_denom if r_denom else None
        rent_severe       = r_b50p / r_denom if r_denom else None
        o_total  = fv("B25091_002E") or 0
        o_notc   = fv("B25091_010E") or 0
        o_b30_34 = fv("B25091_006E") or 0
        o_b35_39 = fv("B25091_007E") or 0
        o_b40_49 = fv("B25091_008E") or 0
        o_b50p   = fv("B25091_009E") or 0
        o_denom  = max(0, o_total - o_notc)
        owner_burdened    = (o_b30_34 + o_b35_39 + o_b40_49 + o_b50p) / o_denom if o_denom else None
        owner_severe      = o_b50p / o_denom if o_denom else None
        return {
            "tenantGRAPI":             graphi,
            "ownerSMOCAPI":            smocapi,
            "rentBurdenedPct":         rent_burdened,
            "rentSeverelyBurdenedPct": rent_severe,
            "ownerBurdenedPct":        owner_burdened,
            "ownerSeverelyBurdenedPct":owner_severe,
        }

    result = {"_year": CENSUS_ACS_YEAR}
    s_hdr, s_row = state_data[0], state_data[1]
    result["State"] = _build_row(s_hdr, s_row)

    c_hdr, *c_rows = county_data
    for row in c_rows:
        key = CENSUS_NAME_MAP.get(row[c_hdr.index("NAME")])
        if key:
            result[key] = _build_row(c_hdr, row)
    return result


def fetch_bls_series_ratio(series_id: str, anchor_year: str) -> tuple[float, str] | tuple[None, None]:
    """
    Generic BLS series ratio fetcher. Returns (latest / anchor_year_avg, latest_period).
    Anchor avg is the mean of all monthly observations in `anchor_year`.
    Used to drive realized-burden nowcasts from the ACS 5-year mid-point
    to the current period for both CPI All Items (owner side) and the
    Hawaiʻi private wages series (income denominator on both sides).
    """
    import json
    api_key = os.environ.get("BLS_API_KEY", "")
    end_year = str(int(anchor_year) + 5)  # cover anchor + ~5 years of nowcast room
    payload = json.dumps({
        "seriesid":  [series_id],
        "startyear": anchor_year,
        "endyear":   end_year,
        **({"registrationkey": api_key} if api_key else {}),
    }).encode()
    try:
        resp = json.loads(fetch_bytes(BLS_API_URL, data=payload, headers={"Content-Type": "application/json"}))
    except (OSError, json.JSONDecodeError) as e:
        print(f"  WARNING: BLS fetch for {series_id} failed ({e})")
        return None, None
    if resp.get("status") != "REQUEST_SUCCEEDED":
        print(f"  WARNING: BLS {series_id} returned {resp.get('status')}: {resp.get('message')}")
        return None, None
    rows = resp["Results"]["series"][0]["data"]
    if not rows:
        return None, None
    # Filter to numeric monthly observations. BLS uses "-" for missing
    # values; skip those rather than crashing the entire nowcast.
    def _parse(r):
        try:
            return float(r["value"])
        except (ValueError, TypeError):
            return None
    monthly = [(r["year"], r["period"], _parse(r))
               for r in rows
               if r["period"].startswith("M") and r["period"] != "M13"]
    monthly = [(y, p, v) for (y, p, v) in monthly if v is not None]
    if not monthly:
        return None, None
    monthly.sort()
    anchor_vals = [v for (y, p, v) in monthly if y == anchor_year]
    if not anchor_vals:
        return None, None
    anchor_avg = sum(anchor_vals) / len(anchor_vals)
    y_last, p_last, v_last = monthly[-1]
    latest_period = f"{y_last}-{p_last[1:]}"   # M04 → 04
    return v_last / anchor_avg, latest_period


def nowcast_burden_anchors(anchor: dict, anchor_year: str) -> dict:
    """
    Apply BLS-driven nowcast factors to the ACS realized-burden anchor.

    Math (per METHODOLOGY.md §2.4):
        rent_factor   = CPI_rent_HNL(latest)  / CPI_rent_HNL(anchor_avg)
        cost_factor   = CPI_all_HNL(latest)   / CPI_all_HNL(anchor_avg)
        income_factor = wage_HI(latest)       / wage_HI(anchor_avg)

        tenantRentPTI    = tenantGRAPI  × rent_factor / income_factor
        mortgageOwnerPTI = ownerSMOCAPI × cost_factor / income_factor

    Factors apply uniformly across counties (BLS doesn't publish
    county-level CPI or wages for Hawaiʻi; statewide is the finest
    Hawaiʻi-specific series available). Cost-burden shares are not
    nowcasted — they're tagged as ACS 5-yr in the UI.

    Returns anchor extended with nowcasted fields per county.
    """
    if not anchor:
        return {}

    # Cost-burden shares are NOT nowcasted (they're slow-moving ACS 5-yr
    # distributional shares). Round them up-front so the field values are
    # tight regardless of whether the BLS-driven PTI nowcast succeeds —
    # otherwise a transient BLS API outage would leak raw floats with
    # 16 decimal digits into the HTML countyData block.
    for key in ("State", "Honolulu", "Maui", "Hawaii", "Kauai"):
        if key not in anchor:
            continue
        a = anchor[key]
        for k in ("rentBurdenedPct", "rentSeverelyBurdenedPct",
                  "ownerBurdenedPct", "ownerSeverelyBurdenedPct"):
            if a.get(k) is not None:
                a[k] = round(a[k], 4)

    rent_ratio,  _ = fetch_bls_series_ratio(BLS_RENT_SERIES,       anchor_year)
    cost_ratio,  _ = fetch_bls_series_ratio(BLS_CPI_ALL_ITEMS_HNL, anchor_year)
    wage_ratio,  nowcast_period = fetch_bls_series_ratio(BLS_WAGES_HI_PRIVATE, anchor_year)
    if None in (rent_ratio, cost_ratio, wage_ratio):
        print(f"  WARNING: realized-burden nowcast skipped (missing BLS factor: "
              f"rent={rent_ratio}, cost={cost_ratio}, wage={wage_ratio})")
        return anchor
    print(f"  Nowcast factors ({anchor_year} → {nowcast_period}): "
          f"rent={rent_ratio:.3f}, cost={cost_ratio:.3f}, wage={wage_ratio:.3f}")
    for key in ("State", "Honolulu", "Maui", "Hawaii", "Kauai"):
        if key not in anchor:
            continue
        a = anchor[key]
        if a.get("tenantGRAPI") is not None:
            a["tenantRentPTI"] = round(a["tenantGRAPI"] * (rent_ratio / wage_ratio), 4)
        if a.get("ownerSMOCAPI") is not None:
            a["mortgageOwnerPTI"] = round(a["ownerSMOCAPI"] * (cost_ratio / wage_ratio), 4)
    anchor["_nowcastPeriod"] = nowcast_period
    return anchor


def fetch_bls_rent_ratio() -> tuple[float, float, str, float | None]:
    """
    Fetch BLS CPI series CUURS49ASEHA (Honolulu MSA, rent of primary residence)
    and return (ratio_now, ratio_smoothed, period, yoy_pct).

    ratio_now     — latest_idx / RENT_ANCHOR_YEAR annual avg (single-month)
    ratio_smoothed— mean of the last BLS_RENT_SMOOTHING_WINDOW monthly ratios
                    against the same anchor average. Use this for the blended
                    nowcast so a single bumpy CPI print doesn't swing the
                    headline rent (Q1 of the rent-nowcast improvement plan).
                    ZORI input already uses the same window — matching it
                    means both legs of the blend are 3-month means.
    period        — ISO "YYYY-MM" of the latest observation
    yoy_pct       — 12-month YoY % change for the same month one year prior,
                    or None if the prior observation is unavailable.

    Raises on network/parse failure so callers can decide whether to fall back
    to raw ACS values.
    """
    import json
    import datetime

    api_key = os.environ.get("BLS_API_KEY", "")
    current_year = str(datetime.date.today().year)
    payload = json.dumps({
        "seriesid": [BLS_RENT_SERIES],
        "startyear": BLS_BASE_YEAR,
        "endyear": current_year,
        **({"registrationkey": api_key} if api_key else {}),
    }).encode()
    data = json.loads(
        fetch_bytes(BLS_API_URL, data=payload, headers={"Content-Type": "application/json"})
    )

    series_data = data["Results"]["series"][0]["data"]

    # Base-year annual average (exclude M13 annual row, skip "-" missing)
    base_vals = [
        float(r["value"])
        for r in series_data
        if r["year"] == BLS_BASE_YEAR
        and r["period"].startswith("M")
        and r["period"] != "M13"
        and r["value"] != "-"
    ]
    if not base_vals:
        raise ValueError(f"No BLS monthly data found for base year {BLS_BASE_YEAR}")
    base_avg = sum(base_vals) / len(base_vals)

    # Monthly observations newest-first, skipping the annual M13 row and "-".
    monthly = [
        r for r in series_data
        if r["period"].startswith("M")
        and r["period"] != "M13"
        and r["value"] != "-"
    ]
    if not monthly:
        raise ValueError("No recent BLS monthly value found")
    recent = monthly[0]

    current_idx = float(recent["value"])
    ratio_now   = current_idx / base_avg
    period      = f"{recent['year']}-{recent['period'][1:].zfill(2)}"  # e.g. "2026-03"

    # Smoothed ratio — trailing mean of the most recent
    # BLS_RENT_SMOOTHING_WINDOW monthly ratios. Mirrors the ZORI smoothing
    # in fetch_zori_asking_rents so the blended output isn't whipsawed by
    # a single bumpy CPI print (BLS Honolulu rent has notoriously sparse
    # sampling — each unit gets revisited every 6 months, so single-print
    # noise can be substantial).
    window = monthly[:BLS_RENT_SMOOTHING_WINDOW]
    ratio_smoothed = sum(float(r["value"]) / base_avg for r in window) / len(window)

    # YoY vs. same month a year ago (skip if prior-year observation missing —
    # common when the "current year" is also the anchor year).
    prior_year = str(int(recent["year"]) - 1)
    prior = next(
        (
            r for r in series_data
            if r["year"] == prior_year
            and r["period"] == recent["period"]
            and r["value"] != "-"
        ),
        None,
    )
    yoy_pct: float | None = None
    if prior is not None:
        prior_val = float(prior["value"])
        if prior_val > 0:
            yoy_pct = (current_idx / prior_val - 1.0) * 100.0

    yoy_str = f", YoY {yoy_pct:+.2f}%" if yoy_pct is not None else ""
    print(f"  BLS {BLS_RENT_SERIES}: base_avg={base_avg:.2f}, current={current_idx:.3f}, "
          f"ratio={ratio_now:.4f} (3-mo smoothed {ratio_smoothed:.4f}, period {period}{yoy_str})")
    return ratio_now, ratio_smoothed, period, yoy_pct


def fetch_bls_rent(honolulu_acs_anchor: int) -> dict:
    """
    Scales the live ACS Honolulu contract rent (from the current run's
    fetch_census_rent()) by the BLS CPI index ratio to produce a
    monthly-current estimate for existing-tenant rent in Honolulu.
    Neighbor islands are scaled in main() using the ratio directly
    (see fetch_bls_rent_ratio).

    *honolulu_acs_anchor* must be the ACS {RENT_ANCHOR_YEAR} 5-year
    Honolulu contract rent (dollars). Passing the wrong vintage here
    double-scales the index and produces a silently wrong dollar value.

    Returns {"Honolulu": {"rent": int}, "_period": "YYYY-MM",
             "_ratio": float (latest, single-month),
             "_ratio_smoothed": float (3-month trailing mean for blend),
             "_yoy_pct": float|None}.

    The CPI-only `rent` value uses the latest single-month ratio (it's a
    display fallback when the blend can't run). The blended nowcast in
    _fetch_rents() reads _ratio_smoothed instead.
    """
    ratio_now, ratio_smoothed, period, yoy_pct = fetch_bls_rent_ratio()
    scaled_rent   = round(honolulu_acs_anchor * ratio_now)
    print(f"  → Honolulu rent ${scaled_rent:,} "
          f"(anchor ACS {RENT_ANCHOR_YEAR} ${honolulu_acs_anchor:,} × ratio {ratio_now:.4f}, "
          f"BLS period {period})")
    return {
        "Honolulu":         {"rent": scaled_rent},
        "_period":          period,
        "_ratio":           ratio_now,
        "_ratio_smoothed":  ratio_smoothed,
        "_yoy_pct":         yoy_pct,
    }


def blend_rent_nowcast(
    acs_anchor: float,
    bls_ratio: float,
    zori_ratio: float,
    cpi_weight: float = BLENDED_RENT_CPI_WEIGHT,
    *,
    fmr_ratio:   float | None = None,
    fmr_weight:  float = 0.0,
    zori_weight: float | None = None,
) -> dict:
    """
    Blend BLS-CPI-scaled rent (lagging ~12 mo) with ZORI-implied rent (leading),
    plus an optional HUD-FMR third leg, to nowcast current tenant rent. Every
    component uses the same ACS dollar anchor (RENT_ANCHOR_YEAR); only the
    anchor→present growth factor differs.

      bls_ratio    = CUURS49ASEHA(latest 3-mo trailing mean) / anchor annual avg
                     [smoothed via BLS_RENT_SMOOTHING_WINDOW — see fetch_bls_rent_ratio]
      zori_ratio   = ZORI(latest 3-mo trailing mean) / ZORI(anchor annual avg)
                     [per county; state ratio used as proxy when a county's
                      anchor avg is missing]
      fmr_ratio    = HUD 2-BR FMR(latest FY) / FMR(anchor FY) — a genuinely
                     county-specific (and Kauaʻi-available) growth signal. Only
                     supplied for the counties in BLENDED_RENT_3LEG_WEIGHTS
                     (Hawaiʻi, Kauaʻi); None elsewhere → 2-leg blend.
      cpi_weight   = per-county CPI weight (see BLENDED_RENT_CPI_WEIGHTS /
                     BLENDED_RENT_3LEG_WEIGHTS).

    Weight handling:
      • 2-leg (default): fmr_weight=0, zori_weight=None → zori_weight = 1−cpi_weight.
      • 3-leg: caller passes explicit cpi/zori/fmr weights (summing to 1).
      blended_factor = cpi_w·bls_ratio + zori_w·zori_ratio + fmr_w·fmr_ratio

    If fmr_ratio is None or fmr_weight is 0 the FMR leg is dropped and its
    weight is folded back onto ZORI so the blend never silently under-sums.

    Returns a dict with the blended value plus each single-source component so
    callers can log them all (useful for audit and the methodology tooltip).
    """
    cpi_w = cpi_weight
    use_fmr = fmr_ratio is not None and fmr_weight > 0
    fmr_w   = fmr_weight if use_fmr else 0.0
    # Default ZORI weight makes the 2-leg case behave exactly as before; in the
    # 3-leg case the caller passes an explicit zori_weight. If the FMR leg is
    # requested but unavailable, its weight is reassigned to ZORI.
    if zori_weight is None:
        zori_w = 1.0 - cpi_w
    else:
        zori_w = zori_weight + (fmr_weight if not use_fmr else 0.0)

    blended_factor = cpi_w * bls_ratio + zori_w * zori_ratio
    if use_fmr:
        blended_factor += fmr_w * fmr_ratio

    return {
        "blended":      round(acs_anchor * blended_factor),
        "cpi_scaled":   round(acs_anchor * bls_ratio),
        "zori_implied": round(acs_anchor * zori_ratio),
        "fmr_implied":  round(acs_anchor * fmr_ratio) if use_fmr else None,
        "bls_ratio":    bls_ratio,
        "zori_ratio":   zori_ratio,
        "fmr_ratio":    fmr_ratio if use_fmr else None,
        "cpi_weight":   cpi_w,
        "zori_weight":  zori_w,
        "fmr_weight":   fmr_w,
    }


def _load_ntr_atr_benchmarks() -> dict:
    """Load the manually-refreshed national NTR/ATR YoY benchmark file.

    Missing file or missing values → returns an empty-ish dict with `ntr_yoy_pct`
    and `atr_yoy_pct` both None. Callers degrade gracefully to a reduced audit.
    """
    import json
    empty = {"ntr_yoy_pct": None, "atr_yoy_pct": None,
             "latest_quarter": None, "last_refreshed": None,
             "sanity_band_pp": 8.0}
    if not NTR_ATR_BENCHMARKS_PATH.exists():
        return empty
    try:
        d = json.loads(NTR_ATR_BENCHMARKS_PATH.read_text())
    except (json.JSONDecodeError, OSError):
        return empty
    return {
        "ntr_yoy_pct":    d.get("ntr_yoy_pct"),
        "atr_yoy_pct":    d.get("atr_yoy_pct"),
        "latest_quarter": d.get("latest_quarter"),
        "last_refreshed": d.get("last_refreshed"),
        "sanity_band_pp": float(d.get("_sanity_band_pp") or 8.0),
    }


def audit_rent_nowcast_vs_ntr(
    hnl_cpi_yoy:  float | None,
    hnl_zori_yoy: float | None,
    cpi_weight:   float = BLENDED_RENT_CPI_WEIGHT,
) -> None:
    """Print a dev-facing sanity check: Honolulu rent signals vs national NTR/ATR.

    No HTML surface, no return value. Run after the blended nowcast so we can
    log what percentage change each component (existing tenant, asking, blended)
    is showing and compare it to the BLS national research series. A divergence
    greater than *sanity_band_pp* prints a WARNING — it doesn't block the run.

    BLS R-CPI-NTR/R-CPI-ATR are national only (no Hawaii cut exists), so this
    is a directional check, not an equality test. We expect Hawaii to run
    hotter than national during tight-supply periods and cooler when the
    mainland surges (post-2021).
    """
    bm = _load_ntr_atr_benchmarks()
    ntr, atr = bm["ntr_yoy_pct"], bm["atr_yoy_pct"]

    # Honolulu blended YoY: since blend is linear in the two ratios and both
    # ratios share the same ACS anchor, the blended 12-month % change equals
    # w·CPI_YoY + (1-w)·ZORI_YoY (first-order Taylor; ratios near 1 → the
    # approximation is within ~0.1 pp of the exact blended ratio-of-ratios).
    hnl_blended_yoy = None
    if hnl_cpi_yoy is not None and hnl_zori_yoy is not None:
        hnl_blended_yoy = cpi_weight * hnl_cpi_yoy + (1 - cpi_weight) * hnl_zori_yoy

    band = bm["sanity_band_pp"]
    q    = bm["latest_quarter"] or "— benchmark not yet refreshed —"

    def _fmt(v):
        return f"{v:+.2f}%" if v is not None else "   n/a  "

    print(f"\nRent sanity check — Honolulu vs national NTR/ATR (BLS {q})")
    print(f"  Honolulu BLS rent CPI  YoY: {_fmt(hnl_cpi_yoy)}   "
          f"← existing tenants ({BLS_RENT_SERIES})")
    print(f"  Honolulu ZORI asking   YoY: {_fmt(hnl_zori_yoy)}   "
          f"← new listings (Zillow ZORI)")
    print(f"  National R-CPI-NTR     YoY: {_fmt(ntr)}   "
          f"← new tenants (national research series)")
    print(f"  National R-CPI-ATR     YoY: {_fmt(atr)}   "
          f"← all tenants (national research series)")
    print(f"  Honolulu blended       YoY: {_fmt(hnl_blended_yoy)}   "
          f"← {int(cpi_weight*100)}·CPI + {int((1-cpi_weight)*100)}·ZORI")

    # Warn when Honolulu blended diverges sharply from national NTR — the closest
    # national analog to our nowcast (both emphasize new-tenant momentum). If no
    # benchmark, skip with a hint so the user knows why.
    if ntr is None:
        print(f"  (no NTR benchmark on file — refresh {NTR_ATR_BENCHMARKS_PATH.name} "
              f"from https://www.bls.gov/cpi/research-series/r-cpi-ntr.htm)")
        return
    if hnl_blended_yoy is None:
        return
    gap = hnl_blended_yoy - ntr
    if abs(gap) > band:
        print(f"  ⚠ WARNING: Honolulu blended YoY differs from national NTR by "
              f"{gap:+.2f} pp (band ±{band:.1f} pp). Investigate before publishing.")


def fetch_dbedt_construction() -> dict:
    """
    Download DBEDT QSER construction XLSX and extract E-8:
    'Estimated Value of Private Building Construction Authorizations, By County'
    (in thousands of dollars, quarterly).

    Returns {countyKey: buildAuth_millions} for the most recent complete year,
    plus a '_period' key with the year string.
    Requires openpyxl (pip install openpyxl).
    """
    if not _OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl is required for DBEDT fetch — run: pip install openpyxl")

    print(f"  Downloading E-construction-tables.xlsx...")
    raw = fetch_bytes(DBEDT_URL)

    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb["E-8"]

    rows = list(ws.iter_rows(values_only=True))

    # Locate the header row — it contains "State" in column 1
    header_idx = None
    for i, row in enumerate(rows):
        if row and len(row) > 1 and row[1] is not None and str(row[1]).strip() == "State":
            header_idx = i
            break

    if header_idx is None:
        raise ValueError("Could not find header row in E-8 worksheet")

    # Data columns: 1=State, 2=Honolulu, 3=Hawaii County, 4=Kauai County, 5=Maui County
    # (indices align with DBEDT_COL_KEYS order)
    data_col_indices = [1, 2, 3, 4, 5]

    # Collect annual rows (skip quarterly "Qtr." rows and float-valued % change rows)
    annual_data = {}
    for row in rows[header_idx + 2:]:   # +2 skips header + "In Thousands" label
        if not row or row[0] is None:
            continue
        year_cell = str(row[0]).strip()

        # Skip quarterly rows
        if "Qtr" in year_cell or "qtr" in year_cell:
            continue
        # Skip percentage-change section (first cell is a float)
        if isinstance(row[0], float):
            continue

        # Clean year string: strip "1/  ", "2/  " footnote prefixes
        year_clean = re.sub(r"^\d+/\s*", "", year_cell).strip()
        try:
            year = int(float(year_clean))
        except (ValueError, TypeError):
            continue

        row_vals = {}
        for j, key in zip(data_col_indices, DBEDT_COL_KEYS):
            if j < len(row) and isinstance(row[j], (int, float)):
                row_vals[key] = row[j]  # thousands of dollars

        if row_vals:
            annual_data[year] = row_vals

    if not annual_data:
        raise ValueError("No annual data parsed from E-8 — check sheet structure")

    latest_year = max(annual_data.keys())
    latest = annual_data[latest_year]

    result = {"_period": str(latest_year)}
    for key, val_thousands in latest.items():
        result[key] = round(val_thousands / 1000)  # → millions, rounded

    return result


def fetch_hhfdc_county_mfi() -> dict:
    """
    Download HHFDC county income schedule PDFs and extract HUD median family
    income (4-person) for each county. The MFI appears as the first dollar
    figure on page 1 (e.g. "$129,300" for Honolulu).

    Returns {countyKey: {"income": int}}, plus a "_period" key.
    Requires pdfplumber (pip install pdfplumber).
    """
    if not _PDFPLUMBER_AVAILABLE:
        raise ImportError("pdfplumber is required for HHFDC fetch — run: pip install pdfplumber")

    result = {"_period": HUD_FY}
    for county in HHFDC_COUNTIES:
        url = HHFDC_PDF_TEMPLATE.format(county=county)
        print(f"  Downloading {county}-County-2025.pdf...")
        raw = fetch_bytes(url)

        with pdfplumber.open(io.BytesIO(raw)) as pdf:
            text = pdf.pages[0].extract_text() or ""

        m = re.search(r"\$(\d{2,3}(?:,\d{3})+)", text)
        if not m:
            print(f"  WARNING: could not parse MFI from {county} PDF")
            continue
        result[county] = {"income": int(m.group(1).replace(",", ""))}

    return result


def fetch_hud_state_mfi() -> dict:
    """
    Download HUD's FY 2025 State Income Limits report PDF and extract the
    Hawaii statewide median family income. The Hawaii row appears as:
        HAWAII
        FY 2025 MFI: 123000 30% OF MEDIAN ...

    Returns {"State": {"income": int}, "_period": HUD_FY}.
    Requires pdfplumber.
    """
    if not _PDFPLUMBER_AVAILABLE:
        raise ImportError("pdfplumber is required for HUD state fetch — run: pip install pdfplumber")

    print(f"  Downloading State-Incomelimits-Report-FY25.pdf...")
    raw = fetch_bytes(HUD_STATE_IL_URL)

    with pdfplumber.open(io.BytesIO(raw)) as pdf:
        # Search all pages for Hawaii — alphabetically it's on page 1, but don't hardcode.
        text = "\n".join((p.extract_text() or "") for p in pdf.pages)

    m = re.search(r"HAWAII\s+FY\s*2025\s*MFI:\s*(\d+)", text)
    if not m:
        raise ValueError("Could not parse Hawaii MFI from HUD state PDF")

    return {"_period": HUD_FY, "State": {"income": int(m.group(1))}}


# ─── HUD Fair Market Rent (FMR) — county-specific 3rd blend leg ──────────────
# HUD publishes a single combined historical workbook ("FMR_2Bed_YYYY_YYYY.xlsx")
# with ONE column per fiscal year ("fmrNN_2" = that FY's 2-BR FMR). Using it
# avoids the per-year filename + header drift that plagues the individual FMR
# releases — we only have to resolve one link (its end-year bumps each fall).
# The HUD portal bot-blocks default User-Agents, so we send a browser UA, and
# HUD workbooks ship a malformed dcterms date that openpyxl rejects, so we
# strip it first. See BLENDED_RENT_3LEG_WEIGHTS for the why/which-counties.
HUD_FMR_DATASETS_PAGE     = "https://www.huduser.gov/portal/datasets/fmr.html"
HUD_FMR_PAGE_BASE         = "https://www.huduser.gov/portal/datasets/"
HUD_FMR_COMBINED_RE       = re.compile(r'href="([^"]*FMR_2Bed_\d{4}_\d{4}\.xlsx)"', re.I)
# Last-known-good combined file, used only if the page scrape finds no link.
HUD_FMR_COMBINED_FALLBACK = "https://www.huduser.gov/portal/datasets/FMR/FMR_2Bed_1983_2026.xlsx"
HUD_FMR_BROWSER_UA        = "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36"
HUD_FMR_STATE_FIPS        = "15"   # Hawaii
HUD_FMR_COUNTY_FIPS       = {"001": "Hawaii", "003": "Honolulu", "007": "Kauai", "009": "Maui"}


def _sanitize_hud_xlsx(data: bytes) -> bytes:
    """HUD workbooks ship a malformed <dcterms:created/modified> that openpyxl
    rejects; strip those two elements from docProps/core.xml and re-zip."""
    import zipfile
    bin_in, out = io.BytesIO(data), io.BytesIO()
    with zipfile.ZipFile(bin_in) as zin, zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zout:
        for n in zin.namelist():
            d = zin.read(n)
            if n == "docProps/core.xml":
                t = d.decode("utf-8", "ignore")
                t = re.sub(r"<dcterms:(created|modified)[^>]*>.*?</dcterms:\1>", "", t)
                d = t.encode("utf-8")
            zout.writestr(n, d)
    return out.getvalue()


def _hud_fmr_combined_url() -> str:
    """Resolve the current 'FMR_2Bed_YYYY_YYYY.xlsx' combined-history link from
    the HUD FMR datasets page. Falls back to the last-known-good URL on any
    scrape failure (the filename's end-year bumps each fiscal year)."""
    try:
        html = fetch_bytes(HUD_FMR_DATASETS_PAGE,
                           headers={"User-Agent": HUD_FMR_BROWSER_UA}).decode("utf-8", "ignore")
        m = HUD_FMR_COMBINED_RE.search(html)
        if m:
            href = m.group(1)
            if href.startswith("http"):
                return href
            if href.startswith("/"):
                return "https://www.huduser.gov" + href
            return HUD_FMR_PAGE_BASE + href
    except Exception as e:
        print(f"  WARNING: HUD FMR page scrape failed ({e}) — using last-known URL")
    return HUD_FMR_COMBINED_FALLBACK


def _fmr_col_year(col: str | None) -> int | None:
    """'fmr24_2' → 2024, 'fmr99_2' → 1999. Returns None for non-2BR columns.
    Two-digit year: 00-50 → 20xx, 51-99 → 19xx (file spans 1983→present)."""
    m = re.fullmatch(r"fmr(\d{2})_2", col or "")
    if not m:
        return None
    yy = int(m.group(1))
    return (2000 + yy) if yy <= 50 else (1900 + yy)


def fetch_hud_fmr_ratios(anchor_year: str = RENT_ANCHOR_YEAR) -> dict:
    """Return {countyKey: fmr_growth_ratio} where ratio = 2-BR FMR(latest FY) /
    2-BR FMR(anchor FY), for the Hawaii counties. Plus metadata keys
    '_anchor_fy' and '_latest_fy' (e.g. "FY2024", "FY2026").

    Used as the optional 3rd blend leg for the counties in
    BLENDED_RENT_3LEG_WEIGHTS. Returns {} on ANY failure so the blend
    gracefully degrades to the 2-leg CPI/ZORI nowcast for those counties.
    """
    if not _OPENPYXL_AVAILABLE:
        print("  WARNING: openpyxl unavailable — skipping HUD FMR leg")
        return {}
    try:
        url = _hud_fmr_combined_url()
        raw = _sanitize_hud_xlsx(
            fetch_bytes(url, headers={"User-Agent": HUD_FMR_BROWSER_UA}, timeout=120)
        )
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True)
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        hdr = list(next(it))
        idx = {c: i for i, c in enumerate(hdr)}
        i_state, i_county = idx.get("state"), idx.get("county")
        if i_state is None or i_county is None:
            raise ValueError("FMR workbook missing 'state'/'county' columns")

        # Map every "fmrNN_2" column to its 4-digit year.
        year_cols = {y: i for c, i in idx.items()
                     if (y := _fmr_col_year(c)) is not None}
        anchor_y = int(anchor_year)
        if anchor_y not in year_cols:
            raise ValueError(f"FMR workbook has no fmr{anchor_y % 100:02d}_2 column "
                             f"(anchor {anchor_y})")
        latest_y = max(year_cols)
        ia, il = year_cols[anchor_y], year_cols[latest_y]

        ratios: dict = {}
        for r in it:
            if str(r[i_state]) != HUD_FMR_STATE_FIPS:
                continue
            ckey = HUD_FMR_COUNTY_FIPS.get(str(r[i_county]).zfill(3))
            if not ckey:
                continue
            base, cur = r[ia], r[il]
            if base and cur:
                ratios[ckey] = float(cur) / float(base)
        if not ratios:
            raise ValueError("no Hawaii county rows matched in FMR workbook")
        ratios["_anchor_fy"] = f"FY{anchor_y}"
        ratios["_latest_fy"] = f"FY{latest_y}"
        return ratios
    except Exception as e:
        print(f"  WARNING: HUD FMR fetch failed ({e}) — FMR leg disabled (2-leg fallback)")
        return {}


ZORI_PERIOD_RE = re.compile(
    r"/\* ZORI_PERIOD_START \*/.*?/\* ZORI_PERIOD_END \*/",
    flags=re.DOTALL,
)
BLS_RENT_PERIOD_RE = re.compile(
    r"/\* BLS_RENT_PERIOD_START \*/.*?/\* BLS_RENT_PERIOD_END \*/",
    flags=re.DOTALL,
)
HOUSING_PERIOD_RE = re.compile(
    r"/\* HOUSING_PERIOD_START \*/.*?/\* HOUSING_PERIOD_END \*/",
    flags=re.DOTALL,
)


def patch_periods(html: str, zori_period: str | None, bls_rent_period: str | None,
                  housing_period: str | None = None) -> str:
    """Patch the ZORI_PERIOD, BLS_RENT_PERIOD, and HOUSING_PERIOD marker
    blocks if present. Missing markers are silently skipped."""
    if zori_period and ZORI_PERIOD_RE.search(html):
        block = (
            "/* ZORI_PERIOD_START */\n"
            f'const zoriLatestPeriod = "{zori_period}";\n'
            "/* ZORI_PERIOD_END */"
        )
        html = ZORI_PERIOD_RE.sub(lambda m: block, html, count=1)
    if bls_rent_period and BLS_RENT_PERIOD_RE.search(html):
        block = (
            "/* BLS_RENT_PERIOD_START */\n"
            f'const blsRentLatestPeriod = "{bls_rent_period}";\n'
            "/* BLS_RENT_PERIOD_END */"
        )
        html = BLS_RENT_PERIOD_RE.sub(lambda m: block, html, count=1)
    if housing_period and HOUSING_PERIOD_RE.search(html):
        block = (
            "/* HOUSING_PERIOD_START */\n"
            f'const housingLatestPeriod = "{housing_period}";\n'
            "/* HOUSING_PERIOD_END */"
        )
        html = HOUSING_PERIOD_RE.sub(lambda m: block, html, count=1)
    return html


def patch_html(html: str, prices: dict) -> str:
    """
    Replace per-county fields in the countyData object via line-anchored
    regex. Three field shapes are handled:

      - int_fields   — `field:1234`            (\\d+ matcher)
      - float_fields — `field:0.305`           ([\\d.-]+ matcher; signed)
      - nested_fields — `field:{k1:v1, k2:v2}` (literal-block replace)
        Nested fields write a fresh `{...}` literal containing only the
        sub-keys we control; missing sub-keys render as `null` so the JS
        renderer can fall back gracefully.
    """
    int_fields   = ("sfhPrice", "condoPrice", "rent", "askRent", "income",
                    "sfhGap", "condoGap", "sfhMortgage", "condoMortgage")
    float_fields = ("tenantRentPTI", "mortgageOwnerPTI",
                    "rentBurdenedPct", "rentSeverelyBurdenedPct",
                    "ownerBurdenedPct", "ownerSeverelyBurdenedPct",
                    "zoriYoY", "cpiRentYoY",
                    "sfhIdx", "condoIdx", "sfhPTI", "condoPTI")
    # Each nested_field maps to its ordered sub-keys for deterministic output.
    nested_fields = {
        "bedroomRent": ("br0", "br1", "br2", "br3plus"),
    }

    def _fmt_nested(field: str, val: dict) -> str:
        sub_keys = nested_fields[field]
        parts = []
        for k in sub_keys:
            v = val.get(k) if isinstance(val, dict) else None
            parts.append(f"{k}:{v if v is not None else 'null'}")
        return f"{field}:{{{','.join(parts)}}}"

    for county_key, vals in prices.items():
        # Find the line with this county
        pattern = rf'^(\s*{re.escape(county_key)}:\s*{{[^}}]*)'

        def replacer(match):
            line_text = match.group(1)
            # Integer fields
            for field in int_fields:
                if field in vals:
                    line_text = re.sub(
                        rf'{field}:\d+',
                        f'{field}:{vals[field]}',
                        line_text
                    )
            # Float fields (decimal with optional ".d+")
            for field in float_fields:
                if field in vals:
                    line_text = re.sub(
                        rf'{field}:-?[\d.]+',
                        f'{field}:{vals[field]}',
                        line_text
                    )
            return line_text

        new_html = re.sub(pattern, replacer, html, flags=re.MULTILINE)

        # Check if anything changed by testing each field individually
        for field in int_fields + float_fields:
            if field in vals and f'{field}:{vals[field]}' not in new_html:
                print(f"  WARNING: could not set {county_key}.{field}")

        html = new_html

        # Nested-object fields — re-match the per-county block (now updated
        # with scalar replacements) and swap the entire nested literal.
        for field in nested_fields:
            if field not in vals:
                continue
            replacement = _fmt_nested(field, vals[field])
            nested_re = rf'(^\s*{re.escape(county_key)}:\s*\{{[^}}]*?){field}:\{{[^}}]*\}}'
            html, n = re.subn(
                nested_re,
                lambda m: m.group(1) + replacement,
                html,
                flags=re.MULTILINE,
            )
            if n == 0:
                print(f"  WARNING: could not set {county_key}.{field} "
                      f"(nested field not found in HTML — first run? add the field stub manually)")

    return html


def _fetch_sale_prices() -> dict:
    """Download Redfin state + county TSVs and return merged price dict.

    Returns {countyKey: {sfhPrice, condoPrice, period, ...}} for all
    Hawaii counties plus "State".  Exits the process on total failure
    (no Hawaii data at all is unrecoverable).
    """
    print("Fetching Redfin housing market data...")
    state_rows  = download_tsv(STATE_URL)
    county_rows = download_tsv(COUNTY_URL)
    prices = {
        **extract_hawaii_prices(state_rows,  region_col="STATE_CODE", region_values={"HI": "State"}),
        **extract_hawaii_prices(county_rows, region_col="REGION",     region_values=COUNTY_MAP),
    }
    if not prices:
        print("ERROR: No Hawaii data found in Redfin exports")
        sys.exit(1)
    return prices


def _fetch_rents(all_prices: dict) -> tuple[
    float | None,   # bls_ratio
    str   | None,   # bls_rent_period
    float | None,   # bls_rent_yoy
    str   | None,   # zori_period
    dict,           # zori_yoy_map  {countyKey: YoY%}
]:
    """Fetch Census ACS, BLS rent CPI, and Zillow ZORI; merge into *all_prices*.

    Execution order matters:
      1. Census ACS contract rent → sets the RENT_ANCHOR_YEAR dollar anchor
         for all counties.  Must run before BLS scaling so the anchors are
         captured before BLS overwrites the Honolulu value.
      2. BLS rent CPI → scales the ACS anchor to the current month.
      3. ZORI → adds asking-rent and 2024 baseline; drives the blended nowcast.
      4. Blended nowcast → overwrites CPI-only rent with 70/30 composite.

    Returns metadata consumed by downstream callers (period strings for HTML
    patching, YoY floats for the NTR/ATR audit).  All failures are soft-warned;
    the function never raises.
    """
    # ── 1. Census ACS contract rent ──────────────────────────────────────────
    print("\nFetching Census ACS contract rent (existing leases)...")
    try:
        census_rents = fetch_census_rent()
        acs_year = census_rents.pop("_year", CENSUS_ACS_YEAR)
        for key, vals in census_rents.items():
            all_prices.setdefault(key, {}).update(vals)
        print(f"  Got contract rent (ACS {acs_year}) for: {', '.join(census_rents.keys())}")
    except Exception as e:
        print(f"  WARNING: Census rent fetch failed ({e}) — rent will not be updated")

    # Snapshot ACS anchors BEFORE BLS/ZORI scaling overwrites them.
    # Both BLS ratio and ZORI ratio are applied to the same anchor year so the
    # blended formula is dimensionally consistent.
    acs_rent_anchor    = {k: v["rent"] for k, v in all_prices.items() if "rent" in v}
    honolulu_acs_anchor = acs_rent_anchor.get("Honolulu")

    # ── 2. BLS rent CPI ──────────────────────────────────────────────────────
    bls_rent_period   = None
    bls_ratio         = None    # single-month — display fallback only
    bls_ratio_smooth  = None    # 3-mo trailing mean — used for the blend
    bls_rent_yoy      = None
    print("\nFetching BLS CPI rent index (existing tenants, monthly)...")
    try:
        if honolulu_acs_anchor is None:
            raise RuntimeError(
                f"no ACS {RENT_ANCHOR_YEAR} Honolulu anchor — Census fetch must precede BLS scaling"
            )
        bls_rent = fetch_bls_rent(honolulu_acs_anchor)
        bls_rent_period  = bls_rent.pop("_period",          None)
        bls_ratio        = bls_rent.pop("_ratio",           None)
        bls_ratio_smooth = bls_rent.pop("_ratio_smoothed",  None)
        bls_rent_yoy     = bls_rent.pop("_yoy_pct",         None)
        for key, vals in bls_rent.items():
            all_prices.setdefault(key, {}).update(vals)
        if bls_ratio:
            for key in ("Maui", "Hawaii", "Kauai", "State"):
                if key in all_prices and "rent" in all_prices[key]:
                    anchor = acs_rent_anchor.get(key, all_prices[key]["rent"])
                    all_prices[key]["rent"] = round(anchor * bls_ratio)
                    print(f"  {key}: ACS ${anchor:,} × {bls_ratio:.4f} "
                          f"= ${all_prices[key]['rent']:,} (BLS-scaled)")
        print(f"  Updated rents to BLS-scaled estimate ({bls_rent_period})")
    except Exception as e:
        print(f"  WARNING: BLS rent fetch failed ({e}) — rents stay at raw ACS values")

    # ── 3. Zillow ZORI ───────────────────────────────────────────────────────
    zori_period      = None
    zori_anchor_avg  = {}
    zori_yoy_map     = {}
    print("\nFetching Zillow ZORI asking rent data...")
    try:
        zori_rents       = fetch_zori_asking_rents()
        zori_period      = zori_rents.pop("_period",      None)
        zori_anchor_avg  = zori_rents.pop("_anchor_avg",  {}) or {}
        zori_anchor_year = zori_rents.pop("_anchor_year", RENT_ANCHOR_YEAR)
        zori_yoy_map     = zori_rents.pop("_yoy_pct",     {}) or {}
        zori_rents.pop("_smoothing_window", None)   # metadata only — not a county
        for key, ask_rent in zori_rents.items():
            all_prices.setdefault(key, {})["askRent"] = ask_rent
        print(f"  Got askRent ({zori_period or '?'}) for: {', '.join(zori_rents.keys())}")
    except Exception as e:
        print(f"  WARNING: Zillow ZORI fetch failed ({e}) — askRent will not be updated")
        zori_anchor_year = RENT_ANCHOR_YEAR

    # ── 3.5 HUD Fair Market Rent (county-specific 3rd leg, targeted) ──────────
    # Only Hawaiʻi & Kauaʻi get an FMR leg (see BLENDED_RENT_3LEG_WEIGHTS). The
    # fetch is fully soft-failing: an empty dict here just means those two
    # counties fall back to their 2-leg CPI/ZORI blend below.
    print("\nFetching HUD Fair Market Rent (county-specific blend leg)...")
    fmr_ratios = fetch_hud_fmr_ratios(RENT_ANCHOR_YEAR)
    fmr_window = ""
    if fmr_ratios:
        fmr_window = f"{fmr_ratios.pop('_anchor_fy', '?')}→{fmr_ratios.pop('_latest_fy', '?')}"
        print(f"  HUD FMR 2-BR growth ({fmr_window}): "
              + ", ".join(f"{k} {v:.3f}" for k, v in fmr_ratios.items()))

    # ── 4. Blended rent nowcast ───────────────────────────────────────────────
    # Per-county weights: Honolulu/State 70/30 CPI/ZORI; Maui 50/50 CPI/ZORI;
    # Hawaiʻi & Kauaʻi use a 3-leg CPI/ZORI/FMR blend (BLENDED_RENT_3LEG_WEIGHTS)
    # when their FMR ratio is available, else fall back to 2-leg. Plus a 3-mo
    # trailing-mean BLS ratio so a single bumpy CPI print doesn't whipsaw the
    # headline. Falls back to CPI-only rent when ZORI/BLS inputs are missing.
    blend_bls_ratio = bls_ratio_smooth if bls_ratio_smooth is not None else bls_ratio
    if blend_bls_ratio and zori_anchor_avg:
        zori_ratios: dict[str, float] = {}
        for key in ("Honolulu", "Maui", "Hawaii", "Kauai", "State"):
            cur  = (all_prices.get(key) or {}).get("askRent")
            base = zori_anchor_avg.get(key)
            if cur and base:
                zori_ratios[key] = cur / base
        # Proxy missing-county ZORI ratios from the state-level ratio (same
        # approach as the Honolulu BLS CPI being applied statewide).
        proxy = zori_ratios.get("State")
        if proxy is not None:
            for key in ("Honolulu", "Maui", "Hawaii", "Kauai"):
                if key not in zori_ratios and (all_prices.get(key) or {}).get("askRent"):
                    zori_ratios[key] = proxy
                    print(f"  {key}: no {zori_anchor_year} ZORI baseline — "
                          f"using state ratio {proxy:.4f} as proxy")

        smooth_tag = " (3-mo smoothed)" if bls_ratio_smooth is not None else ""
        print(f"\nComputing blended rent nowcast — per-county weights{smooth_tag}...")
        for key in ("Honolulu", "Maui", "Hawaii", "Kauai", "State"):
            v = all_prices.get(key)
            if not v or key not in acs_rent_anchor or key not in zori_ratios:
                continue
            w3    = BLENDED_RENT_3LEG_WEIGHTS.get(key)
            fmr_r = fmr_ratios.get(key)
            if w3 and fmr_r is not None:
                # 3-leg CPI/ZORI/FMR (Hawaiʻi, Kauaʻi)
                b = blend_rent_nowcast(
                    acs_anchor  = acs_rent_anchor[key],
                    bls_ratio   = blend_bls_ratio,
                    zori_ratio  = zori_ratios[key],
                    cpi_weight  = w3["cpi"],
                    zori_weight = w3["zori"],
                    fmr_ratio   = fmr_r,
                    fmr_weight  = w3["fmr"],
                )
                v["rent"] = b["blended"]
                print(f"  {key:<9}  {int(round(w3['cpi']*100))}/{int(round(w3['zori']*100))}/"
                      f"{int(round(w3['fmr']*100))} CPI/ZORI/FMR  "
                      f"CPI ${b['cpi_scaled']:>5,}  ZORI ${b['zori_implied']:>5,}  "
                      f"FMR ${b['fmr_implied']:>5,}  → blended ${b['blended']:>5,}  "
                      f"(bls={b['bls_ratio']:.3f}, zori={b['zori_ratio']:.3f}, "
                      f"fmr={b['fmr_ratio']:.3f})")
            else:
                # 2-leg CPI/ZORI (Honolulu, State, Maui — and Hawaiʻi/Kauaʻi
                # if the FMR fetch failed this run)
                cpi_w = BLENDED_RENT_CPI_WEIGHTS.get(key, 0.7)
                b = blend_rent_nowcast(
                    acs_anchor = acs_rent_anchor[key],
                    bls_ratio  = blend_bls_ratio,
                    zori_ratio = zori_ratios[key],
                    cpi_weight = cpi_w,
                )
                v["rent"] = b["blended"]
                zw = int(round((1 - cpi_w) * 100))
                cw = int(round(cpi_w * 100))
                fb = "  [FMR unavailable → 2-leg fallback]" if (w3 and fmr_r is None) else ""
                print(f"  {key:<9}  {cw}/{zw} CPI/ZORI  "
                      f"CPI-scaled ${b['cpi_scaled']:>5,}  "
                      f"ZORI-implied ${b['zori_implied']:>5,}  "
                      f"→ blended ${b['blended']:>5,}  "
                      f"(bls_ratio={b['bls_ratio']:.3f}, zori_ratio={b['zori_ratio']:.3f}){fb}")
    else:
        print("  Skipping blended nowcast (missing BLS ratio or ZORI 2024 baseline)")

    # ── 5. Export per-county YoY signals to countyData (Q2) ─────────────────
    # ZORI YoY is per-county; BLS rent CPI is Honolulu-only so we attach it
    # to the State row and Honolulu (where it's genuine) but not outer islands
    # (where it's regional proxy — shown via per-county ZORI instead).
    for key in ("Honolulu", "Maui", "Hawaii", "Kauai", "State"):
        y = zori_yoy_map.get(key)
        if y is not None:
            all_prices.setdefault(key, {})["zoriYoY"] = round(y / 100.0, 4)
    if bls_rent_yoy is not None:
        for key in ("State", "Honolulu"):
            all_prices.setdefault(key, {})["cpiRentYoY"] = round(bls_rent_yoy / 100.0, 4)

    return bls_ratio, bls_rent_period, bls_rent_yoy, zori_period, zori_yoy_map


def _fetch_income_and_construction(all_prices: dict) -> str:
    """Fetch HHFDC county MFI, HUD state MFI, and DBEDT build auth; merge into *all_prices*.

    Returns the DBEDT period string (e.g. "2025") for use in the summary table.
    All three fetches are soft-warned on failure — the function never raises.
    """
    print("\nFetching HHFDC county median family incomes (HUD FY 2025)...")
    try:
        hhfdc_incomes = fetch_hhfdc_county_mfi()
        hhfdc_incomes.pop("_period", None)
        for key, vals in hhfdc_incomes.items():
            all_prices.setdefault(key, {}).update(vals)
        print(f"  Got income for: {', '.join(hhfdc_incomes.keys())}")
    except Exception as e:
        print(f"  WARNING: HHFDC income fetch failed ({e}) — county income will not be updated")

    print("\nFetching HUD state median family income...")
    try:
        state_income = fetch_hud_state_mfi()
        state_income.pop("_period", None)
        for key, vals in state_income.items():
            all_prices.setdefault(key, {}).update(vals)
        print(f"  Got income for: {', '.join(state_income.keys())}")
    except Exception as e:
        print(f"  WARNING: HUD state income fetch failed ({e}) — state income will not be updated")

    build_period = "?"
    print("\nFetching DBEDT construction authorization data...")
    try:
        dbedt_data = fetch_dbedt_construction()
        build_period = dbedt_data.pop("_period", "?")
        for key, build_auth in dbedt_data.items():
            all_prices.setdefault(key, {})["buildAuth"] = build_auth
        print(f"  Got buildAuth ({build_period}) for: {', '.join(dbedt_data.keys())}")
    except Exception as e:
        print(f"  WARNING: DBEDT construction fetch failed ({e}) — buildAuth will not be updated")

    return build_period


def _monthly_pi(price: float, rate_pct: float) -> float:
    """Monthly principal+interest on a 30-yr fixed loan at *rate_pct*, 20% down."""
    loan = price * (1.0 - DOWN_PAYMENT_FRAC)
    r    = rate_pct / 100.0 / 12.0
    n    = MORTGAGE_TERM_MONTHS
    if r == 0:
        return loan / n
    return loan * r * (1 + r) ** n / ((1 + r) ** n - 1)


def _afford_price(income: float, rate_pct: float) -> float:
    """Max purchase price a household at *income* can afford — mirrors the
    dashboard's calcAffordPrice(): 30% of gross income to P&I, 30-yr fixed,
    20% down."""
    monthly_max = income / 12.0 * DTI_FRONT_FRAC
    r = rate_pct / 100.0 / 12.0
    n = MORTGAGE_TERM_MONTHS
    loan_amt = monthly_max * n if r == 0 else monthly_max * (1 - (1 + r) ** -n) / r
    ltv = max(0.001, 1.0 - DOWN_PAYMENT_FRAC)
    return loan_amt / ltv


def compute_derived_affordability(all_prices: dict, rate_pct: float = MORTGAGE_RATE_PCT) -> None:
    """Recompute the price-derived affordability metrics in place so they track
    the (smoothed) Redfin price instead of remaining frozen literals.

    For each county and home type (sfh/condo) writes:
      • {type}Idx      — affordability index, affordPrice / price × 100
                         (100 = exactly affordable at median income; the UI
                         buckets <55 cost-burdened / <80 stretched / ≥80 ok)
      • {type}Gap      — income-dollar gap: extra annual income needed so the
                         household could afford the median price (0 if already
                         affordable). JS reads incomeNeeded = income + gap.
      • {type}Mortgage — monthly P&I at this price (20% down, 30-yr, rate_pct)
      • {type}PTI      — monthly P&I as a share of gross monthly income
                         (price-card diagnostic; not currently surfaced in UI)

    All four use the SAME rate the slider defaults to (MORTGAGE_RATE_PCT) so the
    static card values agree with the live calculator at first paint. Counties
    missing income or a price are skipped (their literals stay put + a warning).
    """
    print(f"\nRecomputing price-derived affordability metrics @ {rate_pct:.2f}% "
          f"(20% down, 30-yr P&I)…")
    pairs = (
        ("sfhPrice",   "sfhIdx",   "sfhGap",   "sfhMortgage",   "sfhPTI"),
        ("condoPrice", "condoIdx", "condoGap", "condoMortgage", "condoPTI"),
    )
    for key in ("State", "Honolulu", "Maui", "Hawaii", "Kauai"):
        v = all_prices.get(key)
        if not v:
            continue
        income = v.get("income")
        if not income:
            print(f"  {key:<9} skipped — no income")
            continue
        afford = _afford_price(income, rate_pct)
        for price_f, idx_f, gap_f, mort_f, pti_f in pairs:
            price = v.get(price_f)
            if not price:
                continue
            idx      = afford / price * 100.0
            mortgage = _monthly_pi(price, rate_pct)
            # incomeNeeded scales linearly with affordPrice, so the income that
            # would make `afford == price` is income × price/afford; the gap is
            # the shortfall (clamped at 0 once the home is affordable).
            income_needed = income * price / afford
            gap = max(0, int(round(income_needed - income)))
            v[idx_f]  = round(idx, 1)
            v[gap_f]  = gap
            v[mort_f] = int(round(mortgage))
            v[pti_f]  = round(mortgage / (income / 12.0), 4)
        print(f"  {key:<9} SFH idx {v.get('sfhIdx'):>5}  gap ${v.get('sfhGap'):>7,}  "
              f"P&I ${v.get('sfhMortgage'):>6,}   |  Condo idx {v.get('condoIdx'):>5}  "
              f"gap ${v.get('condoGap'):>7,}  P&I ${v.get('condoMortgage'):>6,}")


def _print_summary(all_prices: dict, build_period: str) -> None:
    """Print a formatted table of the latest fetched values."""
    print("\nLatest data:\n")
    print(f"  {'County':<12} {'SFH':>12} {'Condo':>12} {'ContractRent':>13} "
          f"{'AskRent':>10} {'BuildAuth($M)':>14}  {'Period'}")
    print(f"  {'─'*12} {'─'*12} {'─'*12} {'─'*13} {'─'*10} {'─'*14}  {'─'*10}")
    for key in ("State", "Honolulu", "Maui", "Hawaii", "Kauai"):
        if key not in all_prices:
            continue
        v         = all_prices[key]
        sfh       = f"${v['sfhPrice']:>10,}"   if "sfhPrice"   in v else f"{'N/A':>11}"
        condo     = f"${v['condoPrice']:>10,}" if "condoPrice" in v else f"{'N/A':>11}"
        crent     = f"${v['rent']:>11,}"       if "rent"       in v else f"{'N/A':>12}"
        askrent   = f"${v['askRent']:>8,}"     if "askRent"    in v else f"{'N/A':>9}"
        buildauth = f"${v['buildAuth']:>11,}M" if "buildAuth"  in v else f"{'N/A':>13}"
        print(f"  {key:<12} {sfh} {condo} {crent} {askrent} {buildauth}  "
              f"{v.get('period', build_period)}")


def _write_html(
    targets: list[Path],
    all_prices: dict,
    zori_period: str | None,
    bls_rent_period: str | None,
    housing_period: str | None,
    dry_run: bool,
) -> None:
    """Patch and write (or dry-run report) both dashboard HTML files."""
    for target in targets:
        if not target.exists():
            print(f"\nSkipping {target.name} — not found")
            continue
        html    = target.read_text(encoding="utf-8")
        patched = patch_html(html, all_prices)
        patched = patch_periods(patched, zori_period, bls_rent_period, housing_period)
        if patched == html:
            print(f"\n{target.name}: no changes needed — prices already current.")
            continue
        if dry_run:
            print(f"\n[dry-run] would patch {target.name}")
        else:
            target.write_text(patched, encoding="utf-8")
            print(f"\nUpdated {target.name} with latest Redfin prices.")


def main():
    dry_run  = "--dry-run" in sys.argv
    file_idx = sys.argv.index("--file") + 1 if "--file" in sys.argv else None
    targets  = [Path(sys.argv[file_idx])] if file_idx else DEFAULT_FILES

    all_prices = _fetch_sale_prices()

    _, bls_rent_period, bls_rent_yoy, zori_period, zori_yoy_map = _fetch_rents(all_prices)

    # ── Realized-burden anchor + nowcast (renter GRAPI, owner SMOCAPI,
    #    cost-burden share). Pulls ACS 5-year anchor, nowcasts to current
    #    period using BLS rent / all-items / Hawaiʻi wages, then merges
    #    into all_prices so patch_html() writes the new fields. Failure
    #    is non-fatal — fields stay at whatever values are currently in
    #    the HTML (last good nowcast).
    print("\nFetching realized-burden anchor (ACS B25071 / B25092 / B25070 / B25091)...")
    burden = fetch_census_burden_anchor()
    if burden:
        burden = nowcast_burden_anchors(burden, str(CENSUS_ACS_YEAR))
        for key in ("State", "Honolulu", "Maui", "Hawaii", "Kauai"):
            row = burden.get(key) or {}
            target = all_prices.setdefault(key, {})
            for field in ("tenantRentPTI", "mortgageOwnerPTI",
                          "rentBurdenedPct", "rentSeverelyBurdenedPct",
                          "ownerBurdenedPct", "ownerSeverelyBurdenedPct"):
                if row.get(field) is not None:
                    target[field] = row[field]
        nowcast_period = burden.get("_nowcastPeriod")
        if nowcast_period:
            print(f"  Realized-burden nowcast period: {nowcast_period}")

    # ── Per-bedroom price tiles (ACS B25068). Soft-warn on failure —
    #    the dashboard's renderBedroomTiles() falls back to "—" when
    #    bedroomRent is missing/null.
    print("\nFetching Census ACS bedroom rent (B25031)...")
    bedroom = fetch_census_bedroom_rent()
    if bedroom:
        bedroom.pop("_year", None)
        for key, vals in bedroom.items():
            all_prices.setdefault(key, {}).update(vals)

    # Dev-facing NTR/ATR sanity check (prints table, warns on large divergence).
    audit_rent_nowcast_vs_ntr(
        hnl_cpi_yoy  = bls_rent_yoy,
        hnl_zori_yoy = zori_yoy_map.get("Honolulu"),
    )

    build_period = _fetch_income_and_construction(all_prices)

    # Recompute price-derived affordability metrics (idx/gap/mortgage/PTI) so
    # they track the freshly smoothed prices + incomes instead of drifting.
    compute_derived_affordability(all_prices, MORTGAGE_RATE_PCT)

    _print_summary(all_prices, build_period)

    if dry_run:
        print("\n--dry-run: no files modified")
        return

    housing_period = (all_prices.get("State", {}).get("period") or "")[:7] or None
    _write_html(targets, all_prices, zori_period, bls_rent_period, housing_period, dry_run)


if __name__ == "__main__":
    main()
