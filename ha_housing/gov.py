"""State/federal document sources: DBEDT XLSX, HHFDC + HUD PDFs, HUD FMR."""
# Extracted from redfin-price-updater.py — see that file's docstring for the
# pipeline overview. Behavior-preserving split; function bodies are unchanged.
import io
import re

from ha_common.http_client import fetch_bytes

from .config import (
    DBEDT_COL_KEYS, DBEDT_URL, HHFDC_COUNTIES, HHFDC_PDF_TEMPLATE,
    HUD_FY, HUD_STATE_IL_URL, RENT_ANCHOR_YEAR,
)

try:
    import openpyxl
    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False

try:
    import pdfplumber
    _PDFPLUMBER_AVAILABLE = True
except ImportError:
    _PDFPLUMBER_AVAILABLE = False


def fetch_dbedt_construction() -> dict:
    """
    Download DBEDT QSER construction XLSX and extract E-8:
    'Estimated Value of Private Building Construction Authorizations, By County'
    (in thousands of dollars, quarterly).

    Returns {countyKey: buildAuth_millions} for the most recent complete year,
    plus a '_period' key with the year string.
    Requires openpyxl (pip install openpyxl).
    """
    if not _OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl is required for DBEDT fetch — run: pip install openpyxl")

    print(f"  Downloading E-construction-tables.xlsx...")
    raw = fetch_bytes(DBEDT_URL)

    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb["E-8"]

    rows = list(ws.iter_rows(values_only=True))

    # Locate the header row — it contains "State" in column 1
    header_idx = None
    for i, row in enumerate(rows):
        if row and len(row) > 1 and row[1] is not None and str(row[1]).strip() == "State":
            header_idx = i
            break

    if header_idx is None:
        raise ValueError("Could not find header row in E-8 worksheet")

    # Data columns: 1=State, 2=Honolulu, 3=Hawaii County, 4=Kauai County, 5=Maui County
    # (indices align with DBEDT_COL_KEYS order)
    data_col_indices = [1, 2, 3, 4, 5]

    # Collect annual rows (skip quarterly "Qtr." rows and float-valued % change rows)
    annual_data = {}
    for row in rows[header_idx + 2:]:   # +2 skips header + "In Thousands" label
        if not row or row[0] is None:
            continue
        year_cell = str(row[0]).strip()

        # Skip quarterly rows
        if "Qtr" in year_cell or "qtr" in year_cell:
            continue
        # Skip percentage-change section (first cell is a float)
        if isinstance(row[0], float):
            continue

        # Clean year string: strip "1/  ", "2/  " footnote prefixes
        year_clean = re.sub(r"^\d+/\s*", "", year_cell).strip()
        try:
            year = int(float(year_clean))
        except (ValueError, TypeError):
            continue

        row_vals = {}
        for j, key in zip(data_col_indices, DBEDT_COL_KEYS):
            if j < len(row) and isinstance(row[j], (int, float)):
                row_vals[key] = row[j]  # thousands of dollars

        if row_vals:
            annual_data[year] = row_vals

    if not annual_data:
        raise ValueError("No annual data parsed from E-8 — check sheet structure")

    latest_year = max(annual_data.keys())
    latest = annual_data[latest_year]

    result = {"_period": str(latest_year)}
    for key, val_thousands in latest.items():
        result[key] = round(val_thousands / 1000)  # → millions, rounded

    return result


def fetch_hhfdc_county_mfi() -> dict:
    """
    Download HHFDC county income schedule PDFs and extract HUD median family
    income (4-person) for each county. The MFI appears as the first dollar
    figure on page 1 (e.g. "$129,300" for Honolulu).

    Returns {countyKey: {"income": int}}, plus a "_period" key.
    Requires pdfplumber (pip install pdfplumber).
    """
    if not _PDFPLUMBER_AVAILABLE:
        raise ImportError("pdfplumber is required for HHFDC fetch — run: pip install pdfplumber")

    result = {"_period": HUD_FY}
    for county in HHFDC_COUNTIES:
        url = HHFDC_PDF_TEMPLATE.format(county=county)
        print(f"  Downloading {county}-County-2025.pdf...")
        raw = fetch_bytes(url)

        with pdfplumber.open(io.BytesIO(raw)) as pdf:
            text = pdf.pages[0].extract_text() or ""

        m = re.search(r"\$(\d{2,3}(?:,\d{3})+)", text)
        if not m:
            print(f"  WARNING: could not parse MFI from {county} PDF")
            continue
        result[county] = {"income": int(m.group(1).replace(",", ""))}

    return result


def fetch_hud_state_mfi() -> dict:
    """
    Download HUD's FY 2025 State Income Limits report PDF and extract the
    Hawaii statewide median family income. The Hawaii row appears as:
        HAWAII
        FY 2025 MFI: 123000 30% OF MEDIAN ...

    Returns {"State": {"income": int}, "_period": HUD_FY}.
    Requires pdfplumber.
    """
    if not _PDFPLUMBER_AVAILABLE:
        raise ImportError("pdfplumber is required for HUD state fetch — run: pip install pdfplumber")

    print(f"  Downloading State-Incomelimits-Report-FY25.pdf...")
    raw = fetch_bytes(HUD_STATE_IL_URL)

    with pdfplumber.open(io.BytesIO(raw)) as pdf:
        # Search all pages for Hawaii — alphabetically it's on page 1, but don't hardcode.
        text = "\n".join((p.extract_text() or "") for p in pdf.pages)

    m = re.search(r"HAWAII\s+FY\s*2025\s*MFI:\s*(\d+)", text)
    if not m:
        raise ValueError("Could not parse Hawaii MFI from HUD state PDF")

    return {"_period": HUD_FY, "State": {"income": int(m.group(1))}}


# ─── HUD Fair Market Rent (FMR) — county-specific 3rd blend leg ──────────────
# HUD publishes a single combined historical workbook ("FMR_2Bed_YYYY_YYYY.xlsx")
# with ONE column per fiscal year ("fmrNN_2" = that FY's 2-BR FMR). Using it
# avoids the per-year filename + header drift that plagues the individual FMR
# releases — we only have to resolve one link (its end-year bumps each fall).
# The HUD portal bot-blocks default User-Agents, so we send a browser UA, and
# HUD workbooks ship a malformed dcterms date that openpyxl rejects, so we
# strip it first. See BLENDED_RENT_3LEG_WEIGHTS for the why/which-counties.
HUD_FMR_DATASETS_PAGE     = "https://www.huduser.gov/portal/datasets/fmr.html"
HUD_FMR_PAGE_BASE         = "https://www.huduser.gov/portal/datasets/"
HUD_FMR_COMBINED_RE       = re.compile(r'href="([^"]*FMR_2Bed_\d{4}_\d{4}\.xlsx)"', re.I)
# Last-known-good combined file, used only if the page scrape finds no link.
HUD_FMR_COMBINED_FALLBACK = "https://www.huduser.gov/portal/datasets/FMR/FMR_2Bed_1983_2026.xlsx"
HUD_FMR_BROWSER_UA        = "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36"
HUD_FMR_STATE_FIPS        = "15"   # Hawaii
HUD_FMR_COUNTY_FIPS       = {"001": "Hawaii", "003": "Honolulu", "007": "Kauai", "009": "Maui"}


def _sanitize_hud_xlsx(data: bytes) -> bytes:
    """HUD workbooks ship a malformed <dcterms:created/modified> that openpyxl
    rejects; strip those two elements from docProps/core.xml and re-zip."""
    import zipfile
    bin_in, out = io.BytesIO(data), io.BytesIO()
    with zipfile.ZipFile(bin_in) as zin, zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zout:
        for n in zin.namelist():
            d = zin.read(n)
            if n == "docProps/core.xml":
                t = d.decode("utf-8", "ignore")
                t = re.sub(r"<dcterms:(created|modified)[^>]*>.*?</dcterms:\1>", "", t)
                d = t.encode("utf-8")
            zout.writestr(n, d)
    return out.getvalue()


def _hud_fmr_combined_url() -> str:
    """Resolve the current 'FMR_2Bed_YYYY_YYYY.xlsx' combined-history link from
    the HUD FMR datasets page. Falls back to the last-known-good URL on any
    scrape failure (the filename's end-year bumps each fiscal year)."""
    try:
        html = fetch_bytes(HUD_FMR_DATASETS_PAGE,
                           headers={"User-Agent": HUD_FMR_BROWSER_UA}).decode("utf-8", "ignore")
        m = HUD_FMR_COMBINED_RE.search(html)
        if m:
            href = m.group(1)
            if href.startswith("http"):
                return href
            if href.startswith("/"):
                return "https://www.huduser.gov" + href
            return HUD_FMR_PAGE_BASE + href
    except Exception as e:
        print(f"  WARNING: HUD FMR page scrape failed ({e}) — using last-known URL")
    return HUD_FMR_COMBINED_FALLBACK


def _fmr_col_year(col: str | None) -> int | None:
    """'fmr24_2' → 2024, 'fmr99_2' → 1999. Returns None for non-2BR columns.
    Two-digit year: 00-50 → 20xx, 51-99 → 19xx (file spans 1983→present)."""
    m = re.fullmatch(r"fmr(\d{2})_2", col or "")
    if not m:
        return None
    yy = int(m.group(1))
    return (2000 + yy) if yy <= 50 else (1900 + yy)


def fetch_hud_fmr_ratios(anchor_year: str = RENT_ANCHOR_YEAR) -> dict:
    """Return {countyKey: fmr_growth_ratio} where ratio = 2-BR FMR(latest FY) /
    2-BR FMR(anchor FY), for the Hawaii counties. Plus metadata keys
    '_anchor_fy' and '_latest_fy' (e.g. "FY2024", "FY2026").

    Used as the optional 3rd blend leg for the counties in
    BLENDED_RENT_3LEG_WEIGHTS. Returns {} on ANY failure so the blend
    gracefully degrades to the 2-leg CPI/ZORI nowcast for those counties.
    """
    if not _OPENPYXL_AVAILABLE:
        print("  WARNING: openpyxl unavailable — skipping HUD FMR leg")
        return {}
    try:
        url = _hud_fmr_combined_url()
        raw = _sanitize_hud_xlsx(
            fetch_bytes(url, headers={"User-Agent": HUD_FMR_BROWSER_UA}, timeout=120)
        )
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True)
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        hdr = list(next(it))
        idx = {c: i for i, c in enumerate(hdr)}
        i_state, i_county = idx.get("state"), idx.get("county")
        if i_state is None or i_county is None:
            raise ValueError("FMR workbook missing 'state'/'county' columns")

        # Map every "fmrNN_2" column to its 4-digit year.
        year_cols = {y: i for c, i in idx.items()
                     if (y := _fmr_col_year(c)) is not None}
        anchor_y = int(anchor_year)
        if anchor_y not in year_cols:
            raise ValueError(f"FMR workbook has no fmr{anchor_y % 100:02d}_2 column "
                             f"(anchor {anchor_y})")
        latest_y = max(year_cols)
        ia, il = year_cols[anchor_y], year_cols[latest_y]

        ratios: dict = {}
        for r in it:
            if str(r[i_state]) != HUD_FMR_STATE_FIPS:
                continue
            ckey = HUD_FMR_COUNTY_FIPS.get(str(r[i_county]).zfill(3))
            if not ckey:
                continue
            base, cur = r[ia], r[il]
            if base and cur:
                ratios[ckey] = float(cur) / float(base)
        if not ratios:
            raise ValueError("no Hawaii county rows matched in FMR workbook")
        ratios["_anchor_fy"] = f"FY{anchor_y}"
        ratios["_latest_fy"] = f"FY{latest_y}"
        return ratios
    except Exception as e:
        print(f"  WARNING: HUD FMR fetch failed ({e}) — FMR leg disabled (2-leg fallback)")
        return {}
